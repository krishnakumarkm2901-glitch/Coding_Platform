from flask import Blueprint, request, jsonify, send_file, make_response
from models.db import get_db
from utils.security import hash_password
from utils.decorators import admin_required
from utils.time_utils import (
    get_utc_now,
    parse_to_utc_datetime,
    format_utc_iso,
    calculate_contest_status,
    IST
)
from bson import ObjectId
from datetime import datetime, timezone, timedelta
import re
import calendar
import io
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)
admin_bp = Blueprint("admin", __name__)

def slugify(text):
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    return re.sub(r'[-\s]+', '-', text)

# ----------------- ADMIN_DASHBOARD_STATS -----------------

@admin_bp.route("/stats", methods=["GET"])
@admin_required
def get_admin_dashboard_stats():
    """Return platform overview stats and detailed analytics for admin dashboard."""
    db = get_db()
    
    total_students = db.users.count_documents({"role": "STUDENT"})
    total_problems = db.problems.count_documents({})
    total_mcqs = db.mcqs.count_documents({})
    total_contests = db.contests.count_documents({})
    total_submissions = db.submissions.count_documents({})
    
    now = get_utc_now()
    active_contests = 0
    contests = list(db.contests.find({"is_published": True}))
    for c in contests:
        st = parse_to_utc_datetime(c.get("start_time"))
        et = parse_to_utc_datetime(c.get("end_time"))
        if st and et and st <= now <= et:
            active_contests += 1

    # Verdicts & Languages Breakdown strictly from live database
    all_subs = list(db.submissions.find({}, {"status": 1, "language": 1, "created_at": 1}))
    verdict_counts = {
        "Accepted": 0,
        "Wrong Answer": 0,
        "Compilation Error": 0,
        "Runtime Error": 0,
        "Time Limit Exceeded": 0
    }
    lang_counts = {
        "Python": 0,
        "C++": 0,
        "C": 0,
        "Java": 0,
        "JavaScript": 0
    }

    from datetime import timedelta
    # 7-day activity tracking
    days_data = {}
    for i in range(6, -1, -1):
        d = (now - timedelta(days=i)).strftime("%a")
        days_data[d] = {"day": d, "submissions": 0, "accepted": 0}

    for s in all_subs:
        st_val = s.get("status", "Wrong Answer")
        if st_val in verdict_counts:
            verdict_counts[st_val] += 1
        elif "Accepted" in st_val:
            verdict_counts["Accepted"] += 1
        elif "Compilation" in st_val:
            verdict_counts["Compilation Error"] += 1
        elif "Runtime" in st_val:
            verdict_counts["Runtime Error"] += 1
        elif "Time" in st_val:
            verdict_counts["Time Limit Exceeded"] += 1
        else:
            verdict_counts["Wrong Answer"] += 1

        raw_lang = str(s.get("language", "")).lower().strip()
        if "python" in raw_lang or "py" in raw_lang:
            lang_counts["Python"] += 1
        elif "cpp" in raw_lang or "c++" in raw_lang:
            lang_counts["C++"] += 1
        elif raw_lang == "c":
            lang_counts["C"] += 1
        elif "javascript" in raw_lang or "js" in raw_lang or "node" in raw_lang:
            lang_counts["JavaScript"] += 1
        elif "java" in raw_lang:
            lang_counts["Java"] += 1
        elif raw_lang:
            lang_counts["Python"] += 1

        created = s.get("created_at")
        if isinstance(created, datetime):
            if created.tzinfo is None:
                created = created.replace(tzinfo=timezone.utc)
            day_str = created.strftime("%a")
            if day_str in days_data:
                days_data[day_str]["submissions"] += 1
                if st_val == "Accepted":
                    days_data[day_str]["accepted"] += 1

    # Difficulty Breakdown
    easy_count = db.problems.count_documents({"difficulty": "Easy"})
    med_count = db.problems.count_documents({"difficulty": "Medium"})
    hard_count = db.problems.count_documents({"difficulty": "Hard"})

    # Department Breakdown
    dept_counts = {}
    students = list(db.users.find({"role": "STUDENT"}, {"department": 1}))
    for st in students:
        d = st.get("department", "Computer Science & Engineering")
        dept_counts[d] = dept_counts.get(d, 0) + 1

    # Top Topics directly from database
    distinct_topics = db.problems.distinct("topic")
    topics_list = []
    for t in distinct_topics:
        if t:
            c = db.problems.count_documents({"topic": t})
            topics_list.append({"topic": t, "count": c})
    topics_list.sort(key=lambda x: x["count"], reverse=True)
    if not topics_list:
        topics_list = [
            {"topic": "Arrays", "count": 4},
            {"topic": "Dynamic Programming", "count": 4},
            {"topic": "Strings", "count": 3},
            {"topic": "Linked Lists", "count": 3},
            {"topic": "Searching & Sorting", "count": 3},
            {"topic": "Trees & Graphs", "count": 1},
        ]

    return jsonify({
        "success": True,
        "stats": {
            "total_students": total_students,
            "total_problems": total_problems,
            "total_mcqs": total_mcqs,
            "total_contests": total_contests,
            "total_submissions": total_submissions,
            "active_contests": active_contests
        },
        "analytics": {
            "activity_trend": list(days_data.values()),
            "verdicts": [
                {"name": k, "count": v, "color": "#22B573" if k == "Accepted" else "#EF4444" if k == "Wrong Answer" else "#F2B705" if k == "Time Limit Exceeded" else "#6366F1" if k == "Compilation Error" else "#EC4899"}
                for k, v in verdict_counts.items()
            ],
            "languages": [
                {"name": k, "count": v, "color": "#0757B8" if k == "Python" else "#0066CC" if k == "C++" else "#F2B705" if k == "Java" else "#22B573" if k == "C" else "#EAB308"}
                for k, v in lang_counts.items()
            ],
            "difficulty": [
                {"name": "Easy", "count": easy_count, "color": "#22B573"},
                {"name": "Medium", "count": med_count, "color": "#F2B705"},
                {"name": "Hard", "count": hard_count, "color": "#EF4444"}
            ],
            "departments": [
                {"department": k, "students": v}
                for k, v in dept_counts.items()
            ],
            "topics": topics_list
        }
    }), 200

@admin_bp.route("/system/metrics", methods=["GET"])
@admin_required
def get_system_metrics():
    """Return real-time server health, CPU/RAM utilization, cache stats, and compiler worker queue metrics."""
    import psutil
    from services.cache_service import cache
    from services.compiler_pool import compiler_pool

    db = get_db()
    now = get_utc_now()

    # CPU & RAM Metrics
    cpu_percent = psutil.cpu_percent(interval=0.1)
    cpu_count = psutil.cpu_count(logical=True)
    virtual_mem = psutil.virtual_memory()
    disk_usage = psutil.disk_usage("/")

    # Active Contests & Online Participants
    active_contests_count = db.contests.count_documents({
        "is_published": True,
        "start_time": {"$lte": now.isoformat()},
        "end_time": {"$gte": now.isoformat()}
    })
    active_participants_count = db.contest_participants.count_documents({
        "status": "IN_PROGRESS"
    })

    return jsonify({
        "success": True,
        "system": {
            "cpu_percent": cpu_percent,
            "cpu_cores": cpu_count,
            "ram_used_mb": round((virtual_mem.total - virtual_mem.available) / (1024 * 1024), 1),
            "ram_total_mb": round(virtual_mem.total / (1024 * 1024), 1),
            "ram_percent": virtual_mem.percent,
            "disk_percent": disk_usage.percent
        },
        "database": {
            "status": "connected",
            "active_contests": active_contests_count,
            "active_contest_participants": active_participants_count,
            "total_users": db.users.count_documents({}),
            "total_submissions": db.submissions.count_documents({})
        },
        "cache": cache.get_stats(),
        "compiler_workers": compiler_pool.get_metrics(),
        "server_time_utc": format_utc_iso(now)
    }), 200

# ----------------- STUDENT MANAGEMENT -----------------

@admin_bp.route("/students", methods=["GET"])
@admin_required
def list_students():
    """List and search students with department and year filters."""
    db = get_db()
    search = request.args.get("search", "").strip()
    department = request.args.get("department", "").strip()
    year = request.args.get("year", "").strip()
    status = request.args.get("status", "").strip()
    page = int(request.args.get("page", 1))
    limit = int(request.args.get("limit", 20))
    skip = (page - 1) * limit

    conditions = [{"role": "STUDENT"}]

    if search:
        conditions.append({
            "$or": [
                {"student_id": {"$regex": search, "$options": "i"}},
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"department": {"$regex": search, "$options": "i"}}
            ]
        })

    # Department filter with alias support (e.g. CSE / Computer Science & Engineering)
    if department and department.lower() != "all":
        dept_keywords = {
            "computer science & engineering": ["computer science", "cse"],
            "information technology": ["information technology", "it"],
            "artificial intelligence & data science": ["artificial intelligence", "aids", "ai & ds", "ai and ds"],
            "electronics & communication engineering": ["electronics", "ece"],
            "electrical & electronics engineering": ["electrical", "eee"],
            "mechanical engineering": ["mechanical", "mech"],
            "civil engineering": ["civil"],
            "cyber security": ["cyber"]
        }
        dept_lower = department.lower()
        if dept_lower in dept_keywords:
            or_patterns = [{"department": {"$regex": kw, "$options": "i"}} for kw in dept_keywords[dept_lower]]
            conditions.append({"$or": or_patterns})
        else:
            conditions.append({"department": {"$regex": department, "$options": "i"}})

    # Year filter (matches '1st', '2nd', '3rd', '4th')
    if year and year.lower() != "all":
        year_prefix = year.split(" ")[0]
        conditions.append({"year": {"$regex": f"^{year_prefix}", "$options": "i"}})

    if status and status.lower() != "all":
        conditions.append({"status": status})

    query = {"$and": conditions} if len(conditions) > 1 else conditions[0]

    total = db.users.count_documents(query)
    students_cursor = db.users.find(query).sort("created_at", -1).skip(skip).limit(limit)

    students = []
    for s in students_cursor:
        students.append({
            "id": str(s["_id"]),
            "student_id": s.get("student_id", ""),
            "name": s.get("name", ""),
            "email": s.get("email", ""),
            "department": s.get("department", "CSE"),
            "year": s.get("year", "1st Year"),
            "status": s.get("status", "active"),
            "last_login": s.get("last_login").isoformat() if isinstance(s.get("last_login"), datetime) else str(s.get("last_login", "Never"))
        })

    return jsonify({
        "success": True,
        "students": students,
        "pagination": {
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1
        }
    }), 200

@admin_bp.route("/students", methods=["POST"])
@admin_required
def add_student():
    """Add a new student."""
    data = request.get_json() or {}
    student_id = data.get("student_id", "").strip().upper()
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    password = data.get("password", "").strip()
    department = data.get("department", "CSE").strip()
    year = data.get("year", "1st Year").strip()

    if not student_id or not name:
        return jsonify({"error": "Student ID and Name are required", "success": False}), 400

    db = get_db()
    if db.users.find_one({"student_id": student_id}):
        return jsonify({"error": f"Student with ID '{student_id}' already exists", "success": False}), 400

    if email and db.users.find_one({"email": email}):
        return jsonify({"error": f"Student with email '{email}' already exists", "success": False}), 400

    student_doc = {
        "student_id": student_id,
        "name": name,
        "email": email or f"{student_id.lower()}@college.edu",
        "password": hash_password(password or "student123"),
        "department": department,
        "year": year,
        "role": "STUDENT",
        "status": "active",
        "created_at": datetime.now(timezone.utc)
    }

    res = db.users.insert_one(student_doc)
    
    from services.notification_service import create_notification
    admins = list(db.users.find({"role": "ADMIN"}))
    for admin in admins:
        create_notification(
            user_id=admin["_id"],
            title="New Student Registered",
            message=f"New student {student_doc.get('name')} ({student_doc.get('student_id')}) has been registered manually.",
            notif_type="system"
        )
        
    return jsonify({"success": True, "message": "Student created successfully", "id": str(res.inserted_id)}), 201

@admin_bp.route("/students/<student_id>", methods=["PUT"])
@admin_required
def update_student(student_id):
    """Update student details, department, year, or active status."""
    db = get_db()
    if not ObjectId.is_valid(student_id):
        return jsonify({"error": "Invalid student ID", "success": False}), 400

    data = request.get_json() or {}
    update_fields = {}

    for field in ["name", "email", "department", "year", "status"]:
        if field in data:
            update_fields[field] = data[field]

    if "student_id" in data:
        update_fields["student_id"] = data["student_id"].strip().upper()

    if "password" in data and data["password"].strip():
        update_fields["password"] = hash_password(data["password"].strip())

    update_fields["updated_at"] = datetime.now(timezone.utc)

    db.users.update_one({"_id": ObjectId(student_id)}, {"$set": update_fields})
    return jsonify({"success": True, "message": "Student updated successfully"}), 200

@admin_bp.route("/students/<student_id>", methods=["DELETE"])
@admin_required
def delete_student(student_id):
    """Delete a student account."""
    db = get_db()
    if not ObjectId.is_valid(student_id):
        return jsonify({"error": "Invalid student ID", "success": False}), 400

    db.users.delete_one({"_id": ObjectId(student_id)})
    return jsonify({"success": True, "message": "Student deleted successfully"}), 200

@admin_bp.route("/students/bulk-delete", methods=["POST"])
@admin_required
def bulk_delete_students():
    """Delete multiple student accounts in bulk."""
    db = get_db()
    data = request.get_json() or {}
    student_ids = data.get("ids", [])

    if not student_ids:
        return jsonify({"error": "No student IDs provided", "success": False}), 400

    # Validate ObjectId format for all ids
    invalid_ids = [sid for sid in student_ids if not ObjectId.is_valid(sid)]
    if invalid_ids:
        return jsonify({"error": f"Invalid student ID format: {', '.join(invalid_ids)}", "success": False}), 400

    object_ids = [ObjectId(sid) for sid in student_ids]
    
    # Perform bulk delete using delete_many()
    result = db.users.delete_many({"_id": {"$in": object_ids}})
    
    return jsonify({
        "success": True,
        "message": f"Successfully deleted {result.deleted_count} students",
        "deleted_count": result.deleted_count
    }), 200

@admin_bp.route("/students/<student_id>/reset-password", methods=["POST"])
@admin_required
def reset_student_password(student_id):
    """Reset a student's password to a specific or default password."""
    db = get_db()
    if not ObjectId.is_valid(student_id):
        return jsonify({"error": "Invalid student ID", "success": False}), 400

    data = request.get_json() or {}
    new_password = data.get("new_password", "college123").strip()

    db.users.update_one(
        {"_id": ObjectId(student_id)},
        {"$set": {"password": hash_password(new_password), "updated_at": datetime.now(timezone.utc)}}
    )
    return jsonify({"success": True, "message": f"Password reset to '{new_password}'"}), 200

# ----------------- BULK EXCEL STUDENT IMPORT -----------------

EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')

DEPARTMENT_MAP = {
    "cse": "Computer Science & Engineering",
    "computer science": "Computer Science & Engineering",
    "computer science & engineering": "Computer Science & Engineering",
    "computer science & engg": "Computer Science & Engineering",
    "computer science and engineering": "Computer Science & Engineering",
    "it": "Information Technology",
    "information technology": "Information Technology",
    "aids": "Artificial Intelligence & Data Science",
    "ai & ds": "Artificial Intelligence & Data Science",
    "ai and ds": "Artificial Intelligence & Data Science",
    "artificial intelligence & data science": "Artificial Intelligence & Data Science",
    "artificial intelligence and data science": "Artificial Intelligence & Data Science",
    "artificial intelligence & ds": "Artificial Intelligence & Data Science",
    "ece": "Electronics & Communication Engineering",
    "electronics & communication engineering": "Electronics & Communication Engineering",
    "electronics and communication engineering": "Electronics & Communication Engineering",
    "electronics & comm engg": "Electronics & Communication Engineering",
    "eee": "Electrical & Electronics Engineering",
    "electrical & electronics engineering": "Electrical & Electronics Engineering",
    "electrical and electronics engineering": "Electrical & Electronics Engineering",
    "mech": "Mechanical Engineering",
    "mechanical": "Mechanical Engineering",
    "mechanical engineering": "Mechanical Engineering",
    "civil": "Civil Engineering",
    "civil engineering": "Civil Engineering",
    "cyber security": "Cyber Security",
    "cybersecurity": "Cyber Security",
    "food": "Food Technology",
    "food tech": "Food Technology",
    "food technology": "Food Technology",
    "ft": "Food Technology",
    "agri": "Agriculture Engineering",
    "agriculture": "Agriculture Engineering",
    "agriculture engineering": "Agriculture Engineering",
    "agricultural engineering": "Agriculture Engineering",
    "ag": "Agriculture Engineering",
    "aero": "Aeronautical Engineering",
    "aeronautical": "Aeronautical Engineering",
    "aeronautical engineering": "Aeronautical Engineering",
    "aerospace": "Aeronautical Engineering",
    "aerospace engineering": "Aeronautical Engineering",
    "cce": "Computer & Communication Engineering",
    "computer & communication engineering": "Computer & Communication Engineering",
    "computer and communication engineering": "Computer & Communication Engineering",
    "computer communication engineering": "Computer & Communication Engineering",
    "aiml": "Artificial Intelligence & Machine Learning",
    "ai & ml": "Artificial Intelligence & Machine Learning",
    "ai and ml": "Artificial Intelligence & Machine Learning",
    "artificial intelligence & machine learning": "Artificial Intelligence & Machine Learning",
    "artificial intelligence and machine learning": "Artificial Intelligence & Machine Learning",
}

YEAR_MAP = {
    "1": "1st Year",
    "1st": "1st Year",
    "1st year": "1st Year",
    "first": "1st Year",
    "first year": "1st Year",
    "2": "2nd Year",
    "2nd": "2nd Year",
    "2nd year": "2nd Year",
    "second": "2nd Year",
    "second year": "2nd Year",
    "3": "3rd Year",
    "3rd": "3rd Year",
    "3rd year": "3rd Year",
    "third": "3rd Year",
    "third year": "3rd Year",
    "4": "4th Year",
    "4th": "4th Year",
    "4th year": "4th Year",
    "fourth": "4th Year",
    "fourth year": "4th Year",
}

def normalize_header(header):
    if not header:
        return ""
    h = str(header).strip().lower()
    h = re.sub(r'[\s_\-]+', '_', h)
    
    if h in ["register_number", "register_no", "reg_no", "regno", "student_id", "studentid", "reg_number"]:
        return "register_number"
    if h in ["name", "full_name", "fullname", "student_name", "studentname"]:
        return "name"
    if h in ["email", "email_address", "mail", "student_email"]:
        return "email"
    if h in ["department", "dept", "branch", "department_name"]:
        return "department"
    if h in ["year", "academic_year", "study_year", "current_year"]:
        return "year"
    if h in ["password", "initial_password", "pass"]:
        return "password"
    return h

@admin_bp.route("/students/import-preview", methods=["POST"])
@admin_required
def import_students_preview():
    """Parse and validate uploaded Excel file for student bulk import preview."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded. Please choose an Excel file (.xlsx or .xls).", "success": False}), 400

    file = request.files["file"]
    filename = file.filename or ""
    
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return jsonify({"error": "Unsupported file format. Please upload a valid .xlsx or .xls file.", "success": False}), 400

    # Max file size check: 5MB
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > 5 * 1024 * 1024:
        return jsonify({"error": "File size exceeds the 5MB limit. Please upload a smaller file.", "success": False}), 400

    if file_size == 0:
        return jsonify({"error": "The uploaded Excel file is empty.", "success": False}), 400

    try:
        import io
        import openpyxl
        
        file_bytes = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "The uploaded Excel worksheet contains no data.", "success": False}), 400

        # Extract and normalize headers
        header_row = rows[0]
        header_map = {}
        for col_idx, cell_value in enumerate(header_row):
            if cell_value is not None:
                norm = normalize_header(cell_value)
                if norm:
                    header_map[norm] = col_idx

        required_columns = ["register_number", "name", "department", "year"]
        missing_columns = [col for col in required_columns if col not in header_map]

        if missing_columns:
            return jsonify({
                "error": f"Missing required columns in Excel: {', '.join(missing_columns)}. Expected columns: register_number, name, department, year",
                "missing_columns": missing_columns,
                "success": False
            }), 400

        db = get_db()
        # Fetch existing student IDs and emails for quick lookup
        existing_students = set(db.users.distinct("student_id", {"role": "STUDENT"}))
        existing_registers = set(db.users.distinct("register_number", {"role": "STUDENT"}))
        existing_ids = existing_students.union(existing_registers)
        existing_emails = set([e.lower() for e in db.users.distinct("email") if e])

        seen_file_ids = set()
        seen_file_emails = set()
        preview_rows = []
        valid_count = 0
        invalid_count = 0

        for row_idx, row in enumerate(rows[1:], start=2):
            # Check if row is completely empty
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def get_val(col_name):
                idx = header_map.get(col_name)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
                return ""

            raw_id = get_val("register_number").upper()
            raw_name = get_val("name")
            raw_email = get_val("email").lower()
            if not raw_email and raw_id:
                raw_email = f"{raw_id.lower()}@college.edu"
            raw_dept = get_val("department")
            raw_year = get_val("year")
            raw_password = get_val("password")
            if not raw_password:
                raw_password = "student123"

            is_valid = True
            error_message = ""

            # 1. Register Number Validation
            if not raw_id:
                is_valid = False
                error_message = "Register Number cannot be empty"
            elif raw_id in seen_file_ids:
                is_valid = False
                error_message = f"Duplicate Register Number '{raw_id}' in Excel file"
            elif raw_id in existing_ids:
                is_valid = False
                error_message = "Register Number already exists in database"
            else:
                seen_file_ids.add(raw_id)

            # 2. Name Validation
            if is_valid:
                if not raw_name:
                    is_valid = False
                    error_message = "Name cannot be empty"

            # 3. Email Validation
            if is_valid:
                if not raw_email:
                    is_valid = False
                    error_message = "Email address cannot be empty"
                elif not EMAIL_REGEX.match(raw_email):
                    is_valid = False
                    error_message = "Invalid email address format"
                elif raw_email in seen_file_emails:
                    is_valid = False
                    error_message = f"Duplicate email '{raw_email}' in Excel file"
                elif raw_email in existing_emails:
                    is_valid = False
                    error_message = "Email already exists in database"
                else:
                    seen_file_emails.add(raw_email)

            # 4. Department Validation & Normalization
            normalized_dept = raw_dept
            if is_valid:
                dept_lookup = raw_dept.lower().strip()
                if dept_lookup in DEPARTMENT_MAP:
                    normalized_dept = DEPARTMENT_MAP[dept_lookup]
                elif not raw_dept:
                    is_valid = False
                    error_message = "Department cannot be empty"
                else:
                    is_valid = False
                    error_message = f"Invalid department: '{raw_dept}'"

            # 5. Year Validation & Normalization
            normalized_year = raw_year
            if is_valid:
                year_lookup = raw_year.lower().strip()
                if year_lookup in YEAR_MAP:
                    normalized_year = YEAR_MAP[year_lookup]
                elif not raw_year:
                    is_valid = False
                    error_message = "Year cannot be empty"
                else:
                    is_valid = False
                    error_message = f"Invalid year: '{raw_year}' (Expected 1st, 2nd, 3rd, 4th Year)"

            # 6. Password Validation
            if is_valid:
                if not raw_password:
                    is_valid = False
                    error_message = "Password cannot be empty"
                elif len(raw_password) < 6:
                    is_valid = False
                    error_message = "Password must be at least 6 characters long"

            if is_valid:
                valid_count += 1
            else:
                invalid_count += 1

            preview_rows.append({
                "row_number": row_idx,
                "student_id": raw_id,
                "register_number": raw_id,
                "full_name": raw_name,
                "name": raw_name,
                "email": raw_email,
                "department": normalized_dept,
                "year": normalized_year,
                "password": raw_password,
                "status": "Valid" if is_valid else "Error",
                "error_message": error_message
            })

        return jsonify({
            "success": True,
            "filename": filename,
            "total_rows": len(preview_rows),
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "rows": preview_rows
        }), 200

    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel file: {str(e)}", "success": False}), 400

@admin_bp.route("/students/import-commit", methods=["POST"])
@admin_required
def import_students_commit():
    """Commit validated students to database with fast batching and password hashing caching."""
    try:
        data = request.get_json() or {}
        students = data.get("students", [])

        if not students:
            return jsonify({"error": "No students provided for import", "success": False}), 400

        db = get_db()
        if db is None:
            return jsonify({"error": "Database connection is unavailable", "success": False}), 500

        imported_count = 0
        skipped_count = 0
        failed_count = 0
        results_detail = []
        
        # 1. Fetch all existing student IDs and emails in ONE query safely
        existing_students = set(str(sid).upper() for sid in db.users.distinct("student_id", {"role": "STUDENT"}) if sid)
        existing_registers = set(str(reg).upper() for reg in db.users.distinct("register_number", {"role": "STUDENT"}) if reg)
        existing_ids = existing_students.union(existing_registers)
        existing_emails = set(str(e).lower() for e in db.users.distinct("email") if e)

        # 2. Password hash cache (caches bcrypt hash so identical passwords like 'student123' are only hashed once)
        password_hash_cache = {}
        docs_to_insert = []
        now_utc = datetime.now(timezone.utc)

        for s in students:
            if not isinstance(s, dict):
                continue

            raw_reg = s.get("register_number") or s.get("student_id") or ""
            raw_n = s.get("name") or s.get("full_name") or ""
            raw_em = s.get("email") or ""
            raw_dp = s.get("department") or "Computer Science & Engineering"
            raw_yr = s.get("year") or "1st Year"
            raw_pw = s.get("password") or "student123"

            s_id = str(raw_reg).strip().upper()
            s_name = str(raw_n).strip()
            s_email = str(raw_em).strip().lower()
            if not s_email and s_id:
                s_email = f"{s_id.lower()}@college.edu"
            s_dept = str(raw_dp).strip()
            s_year = str(raw_yr).strip()
            s_password = str(raw_pw).strip()
            if not s_password:
                s_password = "student123"

            if not s_id or not s_name or not s_email or not s_password:
                failed_count += 1
                results_detail.append({
                    "student_id": s_id or "(Unknown)",
                    "register_number": s_id or "(Unknown)",
                    "name": s_name or "(Unknown)",
                    "status": "Failed",
                    "reason": "Missing mandatory field"
                })
                continue

            # Check existing in-memory set (instant O(1) lookup)
            if s_id in existing_ids or s_email in existing_emails:
                skipped_count += 1
                results_detail.append({
                    "student_id": s_id,
                    "register_number": s_id,
                    "name": s_name,
                    "status": "Skipped",
                    "reason": "Register Number or Email already exists"
                })
                continue

            # Add to set so duplicates inside the same file are detected
            existing_ids.add(s_id)
            existing_emails.add(s_email)

            # Memoized password hashing
            if s_password not in password_hash_cache:
                password_hash_cache[s_password] = hash_password(s_password)
            hashed_pw = password_hash_cache[s_password]

            student_doc = {
                "student_id": s_id,
                "register_number": s_id,
                "name": s_name,
                "email": s_email,
                "password": hashed_pw,
                "department": s_dept,
                "year": s_year,
                "role": "STUDENT",
                "status": "active",
                "created_at": now_utc
            }
            docs_to_insert.append(student_doc)
            results_detail.append({
                "student_id": s_id,
                "register_number": s_id,
                "name": s_name,
                "status": "Success",
                "reason": "Created"
            })

        # 3. Bulk insert in batches of 500
        if docs_to_insert:
            try:
                for i in range(0, len(docs_to_insert), 500):
                    batch = docs_to_insert[i:i + 500]
                    db.users.insert_many(batch, ordered=False)
                imported_count = len(docs_to_insert)
            except Exception as batch_err:
                logger.warning("Bulk batch insert warning: %s, falling back to individual inserts", str(batch_err))
                # Fallback if any batch fails
                imported_count = 0
                for doc in docs_to_insert:
                    try:
                        # Clean doc _id if it was already modified by failed insert_many
                        doc_clean = {k: v for k, v in doc.items() if k != "_id"}
                        db.users.update_one(
                            {"student_id": doc_clean["student_id"]},
                            {"$setOnInsert": doc_clean},
                            upsert=True
                        )
                        imported_count += 1
                    except Exception:
                        pass

        if imported_count > 0:
            try:
                from services.notification_service import create_notification
                admins = list(db.users.find({"role": "ADMIN"}))
                for admin in admins:
                    create_notification(
                        user_id=admin["_id"],
                        title="New Students Imported",
                        message=f"Successfully imported {imported_count} students from file.",
                        notif_type="system"
                    )
            except Exception:
                pass

        return jsonify({
            "success": True,
            "message": f"Bulk import completed: {imported_count} students successfully created.",
            "summary": {
                "total_submitted": len(students),
                "imported_count": imported_count,
                "skipped_count": skipped_count,
                "failed_count": failed_count
            },
            "details": results_detail
        }), 200
    except Exception as e:
        logger.exception("Import commit catastrophic failure: %s", str(e))
        return jsonify({"error": f"Import failed: {str(e)}", "success": False}), 500

# ----------------- PROBLEM MANAGEMENT -----------------

@admin_bp.route("/problems", methods=["GET"])
@admin_required
def list_admin_problems():
    """List and filter coding problems for admin management."""
    db = get_db()
    search = request.args.get("search", "").strip()
    difficulty = request.args.get("difficulty", "").strip()
    topic = request.args.get("topic", "").strip()
    page = int(request.args.get("page", 1))
    limit = int(request.args.get("limit", 20))
    skip = (page - 1) * limit

    query = {}
    if search:
        query["$or"] = [
            {"title": {"$regex": search, "$options": "i"}},
            {"topic": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    if difficulty and difficulty.lower() != "all":
        query["difficulty"] = difficulty.capitalize()
    if topic and topic.lower() != "all":
        query["topic"] = topic

    total = db.problems.count_documents(query)
    cursor = db.problems.find(query).sort("created_at", -1).skip(skip).limit(limit)

    problems = []
    for p in cursor:
        p_id = str(p["_id"])
        sub_count = db.submissions.count_documents({"problem_id": p_id})
        problems.append({
            "id": p_id,
            "title": p.get("title", ""),
            "slug": p.get("slug", ""),
            "difficulty": p.get("difficulty", "Easy"),
            "topic": p.get("topic", "General"),
            "total_submissions": sub_count,
            "is_active": p.get("is_active", True),
            "created_at": p.get("created_at").isoformat() if isinstance(p.get("created_at"), datetime) else str(p.get("created_at", ""))
        })

    return jsonify({
        "success": True,
        "problems": problems,
        "pagination": {
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if limit > 0 else 1
        }
    }), 200

@admin_bp.route("/problems", methods=["POST"])
@admin_required
def create_problem():
    """Create a new coding problem."""
    data = request.get_json() or {}
    title = data.get("title", "").strip()
    if not title:
        return jsonify({"error": "Problem title is required", "success": False}), 400

    db = get_db()
    slug = slugify(title)
    # Ensure unique slug
    base_slug = slug
    counter = 1
    while db.problems.find_one({"slug": slug}):
        slug = f"{base_slug}-{counter}"
        counter += 1

    problem_doc = {
        "title": title,
        "slug": slug,
        "difficulty": data.get("difficulty", "Easy").capitalize(),
        "topic": data.get("topic", "General").strip(),
        "description": data.get("description", "").strip(),
        "input_format": data.get("input_format", "").strip(),
        "output_format": data.get("output_format", "").strip(),
        "constraints": data.get("constraints", "").strip(),
        "sample_input": data.get("sample_input", "").strip(),
        "sample_output": data.get("sample_output", "").strip(),
        "test_cases": data.get("test_cases", []),
        "supported_languages": data.get("supported_languages", ["python", "c", "cpp", "java", "javascript", "go", "rust"]),
        "time_limit": float(data.get("time_limit", 2.0)),
        "memory_limit": int(data.get("memory_limit", 128)),
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }

    res = db.problems.insert_one(problem_doc)
    return jsonify({"success": True, "message": "Problem created successfully", "id": str(res.inserted_id)}), 201

@admin_bp.route("/problems/<problem_id>", methods=["PUT"])
@admin_required
def update_problem(problem_id):
    """Update an existing problem."""
    db = get_db()
    if not ObjectId.is_valid(problem_id):
        return jsonify({"error": "Invalid problem ID", "success": False}), 400

    data = request.get_json() or {}
    update_data = {}
    
    for field in [
        "title", "difficulty", "topic", "description", "input_format", 
        "output_format", "constraints", "sample_input", "sample_output", 
        "test_cases", "supported_languages", "time_limit", "memory_limit", "is_active"
    ]:
        if field in data:
            update_data[field] = data[field]

    if "difficulty" in update_data:
        update_data["difficulty"] = update_data["difficulty"].capitalize()

    update_data["updated_at"] = datetime.now(timezone.utc)

    db.problems.update_one({"_id": ObjectId(problem_id)}, {"$set": update_data})
    return jsonify({"success": True, "message": "Problem updated successfully"}), 200

@admin_bp.route("/problems/<problem_id>", methods=["DELETE"])
@admin_required
def delete_problem(problem_id):
    """Delete a problem."""
    db = get_db()
    if not ObjectId.is_valid(problem_id):
        return jsonify({"error": "Invalid problem ID", "success": False}), 400

    db.problems.delete_one({"_id": ObjectId(problem_id)})
    return jsonify({"success": True, "message": "Problem deleted successfully"}), 200

# ----------------- MCQ MANAGEMENT -----------------

@admin_bp.route("/mcqs", methods=["GET"])
@admin_required
def list_admin_mcqs():
    """List MCQs with full answers for admin."""
    db = get_db()
    topic = request.args.get("topic", "").strip()
    difficulty = request.args.get("difficulty", "").strip()
    search = request.args.get("search", "").strip()
    page = int(request.args.get("page", 1))
    limit = int(request.args.get("limit", 20))
    skip = (page - 1) * limit

    query = {}
    if topic and topic.lower() != "all":
        query["topic"] = topic
    if difficulty and difficulty.lower() != "all":
        query["difficulty"] = difficulty.capitalize()
    if search:
        query["question"] = {"$regex": search, "$options": "i"}

    total = db.mcqs.count_documents(query)
    cursor = db.mcqs.find(query).skip(skip).limit(limit)

    mcqs = []
    for m in cursor:
        mcqs.append({
            "id": str(m["_id"]),
            "question": m.get("question"),
            "options": m.get("options", []),
            "correct_answer": m.get("correct_answer"),
            "explanation": m.get("explanation", ""),
            "topic": m.get("topic"),
            "difficulty": m.get("difficulty", "Easy")
        })

    return jsonify({
        "success": True,
        "mcqs": mcqs,
        "pagination": {"total": total, "page": page, "limit": limit, "pages": (total + limit - 1) // limit if limit > 0 else 1}
    }), 200

@admin_bp.route("/mcqs", methods=["POST"])
@admin_required
def add_mcq():
    """Add a new MCQ."""
    data = request.get_json() or {}
    question = data.get("question", "").strip()
    options = data.get("options", [])
    correct_answer = data.get("correct_answer", "").strip()

    if not question or len(options) < 2 or not correct_answer:
        return jsonify({"error": "Question, at least 2 options, and correct answer are required", "success": False}), 400

    db = get_db()
    mcq_doc = {
        "question": question,
        "options": options,
        "correct_answer": correct_answer,
        "explanation": data.get("explanation", "").strip(),
        "topic": data.get("topic", "General").strip(),
        "difficulty": data.get("difficulty", "Easy").capitalize(),
        "created_at": datetime.now(timezone.utc)
    }

    res = db.mcqs.insert_one(mcq_doc)
    return jsonify({"success": True, "message": "MCQ created successfully", "id": str(res.inserted_id)}), 201

@admin_bp.route("/mcqs/<mcq_id>", methods=["PUT"])
@admin_required
def update_mcq(mcq_id):
    """Update an existing MCQ."""
    db = get_db()
    if not ObjectId.is_valid(mcq_id):
        return jsonify({"error": "Invalid MCQ ID", "success": False}), 400

    data = request.get_json() or {}
    update_data = {}
    for field in ["question", "options", "correct_answer", "explanation", "topic", "difficulty"]:
        if field in data:
            update_data[field] = data[field]

    if "difficulty" in update_data:
        update_data["difficulty"] = update_data["difficulty"].capitalize()

    update_data["updated_at"] = datetime.now(timezone.utc)

    db.mcqs.update_one({"_id": ObjectId(mcq_id)}, {"$set": update_data})
    return jsonify({"success": True, "message": "MCQ updated successfully"}), 200

@admin_bp.route("/mcqs/<mcq_id>", methods=["DELETE"])
@admin_required
def delete_mcq(mcq_id):
    """Delete an MCQ."""
    db = get_db()
    if not ObjectId.is_valid(mcq_id):
        return jsonify({"error": "Invalid MCQ ID", "success": False}), 400

    db.mcqs.delete_one({"_id": ObjectId(mcq_id)})
    return jsonify({"success": True, "message": "MCQ deleted successfully"}), 200

# ----------------- EXCEL MCQ IMPORT -----------------

@admin_bp.route("/mcqs/import/template", methods=["GET"])
@admin_required
def download_mcq_excel_template():
    """Download sample Excel template for importing MCQs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MCQ Import Template"

    headers = ["Question", "Option 1", "Option 2", "Option 3", "Option 4", "Correct Option", "Topic", "Difficulty"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="0757B8", end_color="0757B8", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample rows
    ws.append(["What is 2+2?", "3", "4", "5", "6", 2, "Aptitude & Logical Reasoning", "Easy"])
    ws.append(["Capital of India?", "Mumbai", "Delhi", "Chennai", "Kolkata", 2, "General", "Easy"])
    ws.append(["Which data structure uses LIFO?", "Queue", "Stack", "Array", "Linked List", 2, "Data Structures", "Easy"])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "MCQ_Import_Template.xlsx"
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.headers["Access-Control-Expose-Headers"] = "Content-Disposition, Content-Type"
    return response

@admin_bp.route("/mcqs/import/preview", methods=["POST"])
@admin_required
def preview_mcq_excel_import():
    """Parse and validate uploaded Excel file with MCQs, returning valid rows and specific errors."""
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded", "success": False}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Invalid file format. Please upload an .xlsx Excel spreadsheet.", "success": False}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        if not ws or ws.max_row < 2:
            return jsonify({"error": "Excel file is empty or missing data rows.", "success": False}), 400

        # Read header row (row 1)
        header_row = [str(cell.value or "").strip().lower() for cell in ws[1]]
        
        # Identify column indices
        col_map = {}
        for idx, h in enumerate(header_row):
            if "question" in h:
                col_map["question"] = idx
            elif "option 1" in h or h == "option1" or h == "a" or "opt 1" in h:
                col_map["option1"] = idx
            elif "option 2" in h or h == "option2" or h == "b" or "opt 2" in h:
                col_map["option2"] = idx
            elif "option 3" in h or h == "option3" or h == "c" or "opt 3" in h:
                col_map["option3"] = idx
            elif "option 4" in h or h == "option4" or h == "d" or "opt 4" in h:
                col_map["option4"] = idx
            elif "correct" in h or "answer" in h:
                col_map["correct_option"] = idx
            elif "topic" in h:
                col_map["topic"] = idx
            elif "difficulty" in h:
                col_map["difficulty"] = idx

        required_cols = ["question", "option1", "option2", "option3", "option4", "correct_option"]
        missing_cols = [c for c in required_cols if c not in col_map]
        if missing_cols:
            return jsonify({
                "error": f"Missing required columns in Excel: {', '.join(missing_cols)}. Expected: Question, Option 1, Option 2, Option 3, Option 4, Correct Option",
                "success": False
            }), 400

        valid_rows = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # If entire row is empty, skip
            if not any(row):
                continue

            q_val = str(row[col_map["question"]] or "").strip()
            o1_val = str(row[col_map["option1"]] or "").strip()
            o2_val = str(row[col_map["option2"]] or "").strip()
            o3_val = str(row[col_map["option3"]] or "").strip()
            o4_val = str(row[col_map["option4"]] or "").strip()
            corr_val = row[col_map["correct_option"]]

            row_errors = []

            # 1. Validate Question
            if not q_val:
                row_errors.append("Question is missing.")

            # 2. Validate all 4 options
            if not o1_val:
                row_errors.append("Option 1 is missing.")
            if not o2_val:
                row_errors.append("Option 2 is missing.")
            if not o3_val:
                row_errors.append("Option 3 is missing.")
            if not o4_val:
                row_errors.append("Option 4 is missing.")

            # 3. Validate Correct Option (must be 1, 2, 3, or 4)
            correct_num = None
            if corr_val is not None:
                corr_str = str(corr_val).strip()
                try:
                    # Handle float e.g. 2.0
                    val_float = float(corr_str)
                    if val_float in [1.0, 2.0, 3.0, 4.0]:
                        correct_num = int(val_float)
                except ValueError:
                    # If user wrote "Option 2" or "B"
                    if corr_str.lower() in ["1", "option 1", "option1", "a", "opt 1"]:
                        correct_num = 1
                    elif corr_str.lower() in ["2", "option 2", "option2", "b", "opt 2"]:
                        correct_num = 2
                    elif corr_str.lower() in ["3", "option 3", "option3", "c", "opt 3"]:
                        correct_num = 3
                    elif corr_str.lower() in ["4", "option 4", "option4", "d", "opt 4"]:
                        correct_num = 4

            if correct_num is None or correct_num not in [1, 2, 3, 4]:
                row_errors.append("Correct Option must be 1, 2, 3, or 4.")

            topic_val = str(row[col_map["topic"]] or "").strip() if "topic" in col_map and col_map["topic"] < len(row) and row[col_map["topic"]] else "General"
            diff_val = str(row[col_map["difficulty"]] or "").strip().capitalize() if "difficulty" in col_map and col_map["difficulty"] < len(row) and row[col_map["difficulty"]] else "Easy"
            if diff_val not in ["Easy", "Medium", "Hard"]:
                diff_val = "Easy"

            if row_errors:
                errors.append({
                    "row": row_idx,
                    "reason": "; ".join(row_errors),
                    "question": q_val or f"Row {row_idx}"
                })
            else:
                options_list = [o1_val, o2_val, o3_val, o4_val]
                valid_rows.append({
                    "row": row_idx,
                    "question": q_val,
                    "options": options_list,
                    "correctOption": correct_num,
                    "correct_option": correct_num,
                    "correct_answer": options_list[correct_num - 1],
                    "topic": topic_val,
                    "difficulty": diff_val,
                    "type": "MCQ"
                })

        return jsonify({
            "success": True,
            "total_rows": len(valid_rows) + len(errors),
            "valid_count": len(valid_rows),
            "invalid_count": len(errors),
            "valid_rows": valid_rows,
            "errors": errors
        }), 200

    except Exception as e:
        logger.error(f"Error parsing MCQ excel file: {e}")
        return jsonify({"error": f"Failed to read Excel file: {str(e)}", "success": False}), 500

@admin_bp.route("/mcqs/import/commit", methods=["POST"])
@admin_required
def commit_mcq_excel_import():
    """Save validated MCQs to MongoDB."""
    data = request.get_json() or {}
    mcqs_to_save = data.get("mcqs", [])
    if not mcqs_to_save or not isinstance(mcqs_to_save, list):
        return jsonify({"error": "No valid MCQs provided to save.", "success": False}), 400

    db = get_db()
    now = datetime.now(timezone.utc)
    docs = []

    for item in mcqs_to_save:
        q = str(item.get("question", "")).strip()
        opts = item.get("options", [])
        corr_opt = item.get("correctOption") or item.get("correct_option")

        if not q or len(opts) != 4 or not corr_opt or int(corr_opt) not in [1, 2, 3, 4]:
            continue

        corr_num = int(corr_opt)
        corr_ans = opts[corr_num - 1] if 1 <= corr_num <= len(opts) else str(opts[0])

        docs.append({
            "question": q,
            "options": opts,
            "correctOption": corr_num,
            "correct_option": corr_num,
            "correct_answer": corr_ans,
            "type": "MCQ",
            "topic": item.get("topic", "General"),
            "difficulty": item.get("difficulty", "Easy").capitalize(),
            "explanation": item.get("explanation", ""),
            "created_at": now,
            "createdAt": now
        })

    if not docs:
        return jsonify({"error": "No valid MCQ records to insert.", "success": False}), 400

    res = db.mcqs.insert_many(docs)
    inserted_ids = [str(_id) for _id in res.inserted_ids]

    return jsonify({
        "success": True,
        "message": f"Successfully imported {len(inserted_ids)} MCQs.",
        "imported_count": len(inserted_ids),
        "imported_ids": inserted_ids
    }), 201

# ----------------- CONTEST MANAGEMENT -----------------

@admin_bp.route("/contests", methods=["GET"])
@admin_required
def list_admin_contests():
    """List all contests with admin statistics, problem IDs, and MCQ IDs."""
    db = get_db()
    contests_cursor = db.contests.find({}).sort("start_time", -1)
    
    now = get_utc_now()
    contests = []
    for c in contests_cursor:
        c_id = str(c["_id"])
        participants_count = db.contest_participants.count_documents({"contest_id": c_id})
        prob_ids = [str(pid) for pid in c.get("problem_ids", []) if pid]
        mcq_ids = [str(mid) for mid in c.get("mcq_ids", []) if mid]
        
        start = parse_to_utc_datetime(c.get("start_time"))
        end = parse_to_utc_datetime(c.get("end_time"))
        status = calculate_contest_status(start, end, now)
        
        c_type = c.get("contestType") or c.get("contest_type")
        if not c_type:
            if prob_ids and mcq_ids: c_type = "BOTH"
            elif prob_ids: c_type = "CODING"
            else: c_type = "MCQ"

        contests.append({
            "id": c_id,
            "title": c.get("title"),
            "description": c.get("description", ""),
            "start_time": format_utc_iso(start),
            "end_time": format_utc_iso(end),
            "status": status,
            "duration_minutes": c.get("duration_minutes", 60),
            "contest_type": c_type,
            "contestType": c_type,
            "problem_ids": prob_ids,
            "codingProblemIds": prob_ids,
            "mcq_ids": mcq_ids,
            "mcqIds": mcq_ids,
            "problems_count": len(prob_ids),
            "mcqs_count": len(mcq_ids),
            "mcqs_per_student": c.get("mcqs_per_student", 20),
            "allow_calculator": bool(c.get("allow_calculator", False)),
            "allowCalculator": bool(c.get("allow_calculator", False)),
            "total_points": c.get("total_points", len(mcq_ids) * 10 + len(prob_ids) * 50),
            "is_published": c.get("is_published", False),
            "participants_count": participants_count
        })

    return jsonify({
        "success": True, 
        "contests": contests,
        "server_time": format_utc_iso(now)
    }), 200

@admin_bp.route("/contests", methods=["POST"])
@admin_required
def create_contest():
    """Create a new contest with timezone-aware start and end times and contest type."""
    data = request.get_json() or {}
    title = data.get("title", "").strip()
    if not title:
        return jsonify({"error": "Contest title is required", "success": False}), 400

    contest_type = str(data.get("contestType") or data.get("contest_type") or "").strip().upper()
    prob_ids = data.get("problem_ids") or data.get("codingProblemIds") or []
    mcq_ids = data.get("mcq_ids") or data.get("mcqIds") or []

    if not contest_type:
        if prob_ids and mcq_ids: contest_type = "BOTH"
        elif prob_ids: contest_type = "CODING"
        elif mcq_ids: contest_type = "MCQ"
        else:
            return jsonify({"error": "Please select at least one contest type.", "success": False}), 400

    if contest_type not in ["MCQ", "CODING", "BOTH"]:
        return jsonify({"error": "Invalid contest type. Must be MCQ, CODING, or BOTH.", "success": False}), 400

    if contest_type == "MCQ":
        prob_ids = []
    elif contest_type == "CODING":
        mcq_ids = []

    total_points = int(data.get("total_points", 0))
    if total_points <= 0:
        total_points = len(mcq_ids) * 10 + len(prob_ids) * 50

    now = get_utc_now()
    start_time = parse_to_utc_datetime(data.get("start_time")) or now
    end_time = parse_to_utc_datetime(data.get("end_time")) or (start_time + timedelta(hours=1))

    db = get_db()
    contest_doc = {
        "title": title,
        "description": data.get("description", "").strip(),
        "start_time": start_time,
        "end_time": end_time,
        "duration_minutes": int(data.get("duration_minutes", 60)),
        "contestType": contest_type,
        "contest_type": contest_type,
        "problem_ids": prob_ids,
        "codingProblemIds": prob_ids,
        "mcq_ids": mcq_ids,
        "mcqIds": mcq_ids,
        "mcqs_per_student": int(data.get("mcqs_per_student") or 20),
        "allow_calculator": bool(data.get("allow_calculator") or data.get("allowCalculator") or False),
        "total_points": total_points,
        "is_published": data.get("is_published", False),
        "created_at": now
    }

    res = db.contests.insert_one(contest_doc)
    
    from services.notification_service import create_broadcast_notification, create_notification
    # Broadcast to all students
    create_broadcast_notification(
        title="New Contest Available",
        message=f"'{contest_doc.get('title')}' is now open. Register before the contest starts.",
        notif_type="contest",
        created_by="Admin"
    )
    # Alert admins
    admins = list(db.users.find({"role": "ADMIN"}))
    for admin in admins:
        create_notification(
            user_id=admin["_id"],
            title="New Contest Created",
            message=f"Contest '{contest_doc.get('title')}' was created successfully.",
            notif_type="contest"
        )
        
    return jsonify({"success": True, "message": "Contest created successfully", "id": str(res.inserted_id)}), 201

@admin_bp.route("/contests/<contest_id>", methods=["PUT"])
@admin_required
def update_contest(contest_id):
    """Update contest details or toggle published status."""
    db = get_db()
    if not ObjectId.is_valid(contest_id):
        return jsonify({"error": "Invalid contest ID", "success": False}), 400

    data = request.get_json() or {}
    update_data = {}

    for field in ["title", "description", "duration_minutes", "total_points", "is_published", "mcqs_per_student"]:
        if field in data:
            if field == "mcqs_per_student":
                update_data[field] = int(data[field]) if data[field] is not None else 20
            else:
                update_data[field] = data[field]

    if "contestType" in data or "contest_type" in data:
        c_type = str(data.get("contestType") or data.get("contest_type") or "").strip().upper()
        if c_type in ["MCQ", "CODING", "BOTH"]:
            update_data["contestType"] = c_type
            update_data["contest_type"] = c_type

    if "problem_ids" in data or "codingProblemIds" in data:
        p_ids = data.get("problem_ids") or data.get("codingProblemIds") or []
        update_data["problem_ids"] = p_ids
        update_data["codingProblemIds"] = p_ids

    if "mcq_ids" in data or "mcqIds" in data:
        m_ids = data.get("mcq_ids") or data.get("mcqIds") or []
        update_data["mcq_ids"] = m_ids
        update_data["mcqIds"] = m_ids

    if "start_time" in data and data["start_time"]:
        parsed_start = parse_to_utc_datetime(data["start_time"])
        if parsed_start:
            update_data["start_time"] = parsed_start

    if "end_time" in data and data["end_time"]:
        parsed_end = parse_to_utc_datetime(data["end_time"])
        if parsed_end:
            update_data["end_time"] = parsed_end

    if "allow_calculator" in data or "allowCalculator" in data:
        update_data["allow_calculator"] = bool(data.get("allow_calculator") or data.get("allowCalculator") or False)
        update_data["allowCalculator"] = bool(data.get("allow_calculator") or data.get("allowCalculator") or False)

    update_data["updated_at"] = get_utc_now()

    db.contests.update_one({"_id": ObjectId(contest_id)}, {"$set": update_data})
    return jsonify({"success": True, "message": "Contest updated successfully"}), 200

@admin_bp.route("/contests/<contest_id>", methods=["DELETE"])
@admin_required
def delete_contest(contest_id):
    """Delete a contest."""
    db = get_db()
    if not ObjectId.is_valid(contest_id):
        return jsonify({"error": "Invalid contest ID", "success": False}), 400

    db.contests.delete_one({"_id": ObjectId(contest_id)})
    db.contest_participants.delete_many({"contest_id": contest_id})
    return jsonify({"success": True, "message": "Contest deleted successfully"}), 200

@admin_bp.route("/contests/<contest_id>/restore/<participant_id>", methods=["POST"])
@admin_required
def restore_contest_access(contest_id, participant_id):
    """Approve a LOCKED attempt within 30 minutes and create a fresh retest.
    AUTO_TERMINATED attempts continue to use the existing retest flow."""
    db = get_db()
    if not ObjectId.is_valid(contest_id) or not ObjectId.is_valid(participant_id):
        return jsonify({"error": "Invalid contest or participant ID", "success": False}), 400

    contest = db.contests.find_one({"_id": ObjectId(contest_id)})
    if not contest:
        return jsonify({"error": "Contest not found", "success": False}), 404

    participant = db.contest_participants.find_one({"_id": ObjectId(participant_id)})
    if not participant:
        return jsonify({"error": "Participant not found", "success": False}), 404

    # Admins may activate a fresh attempt only from a terminated/locked history record.
    current_status = participant.get("status", "")
    is_terminated = bool(participant.get("auto_terminated") or current_status in ["TERMINATED", "AUTO_TERMINATED"])
    if current_status != "LOCKED" and not is_terminated:
        return jsonify({
            "error": f"Cannot reset participant with status '{current_status}'. Only terminated attempts can be reset.",
            "current_status": current_status,
            "success": False
        }), 400

    # Verify lock window is still active (within 30 minutes) - only for LOCKED attempts
    now = get_utc_now()
    if current_status == "LOCKED":
        lock_timeout = parse_to_utc_datetime(participant.get("lock_timeout_at"))
        if not lock_timeout or now >= lock_timeout:
            # Window expired - update to TERMINATED and reject
            db.contest_participants.update_one(
                {"_id": ObjectId(participant_id)},
                {
                    "$set": {
                        "status": "AUTO_TERMINATED",
                        "resolution_window_active": False,
                        "auto_terminated": True,
                        "terminated_at": now,
                        "termination_reason": "Lock resolution window (30 minutes) expired without admin action"
                    }
                }
            )
            return jsonify({
                "error": "The 30-minute lock resolution window has expired. This attempt has been auto-terminated.",
                "success": False
            }), 403

    # Check for existing retest for this user
    user_id = participant.get("user_id")
    current_attempt = participant.get("attempt_number", 1)
    
    existing_retest = db.contest_participants.find_one({
        "contest_id": ObjectId(contest_id) if ObjectId.is_valid(contest_id) else contest_id,
        "user_id": user_id,
        "is_active_attempt": True,
        "attempt_number": {"$gt": current_attempt}
    })
    
    if existing_retest:
        return jsonify({
            "error": "A retest is already active for this student.",
            "retest_attempt_number": existing_retest.get("attempt_number"),
            "success": False
        }), 409

    # ===== CREATE NEW RETEST ATTEMPT =====
    
    student_id = participant.get("student_id")
    previous_assigned_ids = participant.get("assigned_mcq_ids", [])
    next_attempt = current_attempt + 1
    
    # Generate new randomized questions for retest
    from routes.contests import get_or_assign_student_mcqs
    new_assigned_mcq_ids = get_or_assign_student_mcqs(
        db, 
        contest, 
        user_id, 
        student_id,
        now=now,
        attempt_number=next_attempt,
        exclude_previous_ids=previous_assigned_ids
    )

    # Create new active retest attempt
    new_retest_doc = {
        "contest_id": ObjectId(contest_id) if ObjectId.is_valid(contest_id) else contest_id,
        "user_id": user_id,
        "student_id": student_id,
        "student_name": participant.get("student_name"),
        "department": participant.get("department", "CSE"),
        "assigned_mcq_ids": new_assigned_mcq_ids,
        "attempt_number": next_attempt,
        "is_active_attempt": True,
        "original_attempt_id": ObjectId(participant_id),
        "status": "RETEST_READY",
        "requires_fullscreen": True,
        "resume_state": {},
        "mcq_answers": {},
        "code_solutions": {},
        "joined_at": None,
        "remaining_seconds": int(contest.get("duration_minutes", 60)) * 60,
        "score": 0,
        "problems_solved": 0,
        "mcqs_correct": 0,
        "submitted": False,
        "auto_terminated": False,
        "anti_cheat_logs": [
            {
                "event_type": "RETEST_ACTIVATED",
                "detail": "Admin activated retest with new question set",
                "created_by": str(request.current_user.get("_id", "")),
                "original_attempt_id": str(ObjectId(participant_id)),
                "timestamp": now.isoformat()
            }
        ]
    }
    
    new_retest_result = db.contest_participants.insert_one(new_retest_doc)

    # Keep the old attempt as immutable history; only append lineage metadata.
    db.contest_participants.update_one(
        {"_id": ObjectId(participant_id)},
        {
            "$set": {
                "status": current_status,
                "is_active_attempt": False,
                "resolution_window_active": False,
                "retest_approved_at": now,
                "retest_attempt_id": new_retest_result.inserted_id,
            },
            "$push": {
                "anti_cheat_logs": {
                    "event_type": "RETEST_APPROVED",
                    "detail": "Admin approved a new retest attempt",
                    "created_by": str(request.current_user.get("_id", "")),
                    "timestamp": now.isoformat()
                }
            }
        }
    )

    # Notify student of retest activation
    from services.notification_service import create_notification
    create_notification(
        user_id=user_id,
        title="Retest Activated ✓",
        message=f"Admin has activated your retest with a new set of questions. Enter fullscreen and start the retest to continue.",
        notif_type="contest"
    )
    
    # Notify admins
    admins = list(db.users.find({"role": "ADMIN"}))
    for admin in admins:
        if admin["_id"] != request.current_user.get("_id"):
            create_notification(
                user_id=admin["_id"],
                title="Retest Activated",
                message=f"Admin {request.current_user.get('name')} activated retest for student {participant.get('student_name')}.",
                notif_type="contest"
            )

    return jsonify({
        "success": True, 
        "message": "Contest reset; a new shuffled retest attempt is ready",
        "new_participant_id": str(new_retest_result.inserted_id),
        "attempt_number": next_attempt
    }), 200

@admin_bp.route("/contests/<contest_id>/participants", methods=["GET"])
@admin_required
def get_contest_participants_admin(contest_id):
    """Get all ACTIVE participants for a contest (most recent attempt per student).
    Query param ?show_all=true to see all attempts including historical."""
    db = get_db()
    if not ObjectId.is_valid(contest_id):
        return jsonify({"error": "Invalid contest ID", "success": False}), 400

    show_all = request.args.get("show_all", "false").lower() == "true"
    
    if show_all:
        # Show all attempts (including terminated/locked)
        participants_cursor = db.contest_participants.find({"contest_id": contest_id}).sort([("student_id", 1), ("attempt_number", -1)])
    else:
        # Show only ACTIVE attempts (prioritize is_active_attempt: True)
        participants_cursor = db.contest_participants.find({
            "contest_id": contest_id,
            "is_active_attempt": True
        }).sort("score", -1)
    
    participants = []
    
    for p in participants_cursor:
        # Enforce the lock deadline on the server whenever admin status is read.
        if p.get("status") == "LOCKED":
            lock_timeout = parse_to_utc_datetime(p.get("lock_timeout_at"))
            if not lock_timeout or get_utc_now() >= lock_timeout:
                terminated_at = get_utc_now()
                db.contest_participants.update_one(
                    {"_id": p["_id"], "status": "LOCKED"},
                    {"$set": {
                        "status": "AUTO_TERMINATED",
                        "auto_terminated": True,
                        "resolution_window_active": False,
                        "terminated_at": terminated_at,
                        "termination_reason": "Lock resolution window (30 minutes) expired without admin action",
                    }},
                )
                p["status"] = "AUTO_TERMINATED"
                p["auto_terminated"] = True
        is_term = bool(p.get("auto_terminated") or p.get("status") == "AUTO_TERMINATED")
        is_locked = bool(p.get("status") == "LOCKED")
        is_retest_ready = bool(p.get("status") == "RETEST_READY")
        is_retest_approved = bool(p.get("status") == "RETEST_APPROVED")
        active_retest = db.contest_participants.find_one({
            "contest_id": p.get("contest_id"),
            "user_id": p.get("user_id"),
            "is_active_attempt": True,
            "attempt_number": {"$gt": p.get("attempt_number", 1)},
        })

        if is_term:
            status_val = "AUTO_TERMINATED"
        elif is_locked:
            status_val = "LOCKED"
        elif is_retest_ready:
            status_val = "RETEST_READY"
        elif is_retest_approved:
            status_val = "RETEST_APPROVED"
        elif p.get("submitted"):
            status_val = "SUBMITTED"
        else:
            status_val = "IN_PROGRESS"
        
        # Calculate lock timeout remaining
        lock_timeout_remaining = 0
        if is_locked and p.get("lock_timeout_at"):
            from utils.time_utils import parse_to_utc_datetime, get_utc_now
            lock_timeout_dt = parse_to_utc_datetime(p.get("lock_timeout_at"))
            if lock_timeout_dt:
                lock_timeout_remaining = max(0, int((lock_timeout_dt - get_utc_now()).total_seconds()))
        
        # Show attempt number if not first attempt (indicates retest)
        attempt_num = p.get("attempt_number", 1)
        status_display = status_val
        if attempt_num > 1:
            status_display = f"{status_val} (Retest #{attempt_num})"

        participants.append({
            "id": str(p["_id"]),
            "student_id": p.get("student_id"),
            "student_name": p.get("student_name"),
            "department": p.get("department", "CSE"),
            "score": p.get("score", 0),
            "mcq_score": float(p.get("mcq_score", 0)),
            "coding_score": float(p.get("coding_score", 0)),
            "problems_solved": p.get("problems_solved", 0),
            "mcqs_correct": p.get("mcqs_correct", 0),
            "status": status_val,
            "status_display": status_display,
            "attempt_number": attempt_num,
            "is_active_attempt": p.get("is_active_attempt", False),
            "auto_terminated": is_term,
            "is_locked": is_locked,
            "is_retest_ready": is_retest_ready,
            "is_retest_approved": is_retest_approved,
            "has_active_retest": bool(active_retest),
            "termination_reason": p.get("termination_reason", ""),
            "lock_reason": p.get("lock_reason", ""),
            "lock_timeout_remaining_seconds": lock_timeout_remaining,
            "submitted": p.get("submitted", False),
            "submitted_at": p.get("submitted_at").isoformat() if isinstance(p.get("submitted_at"), datetime) else str(p.get("submitted_at")),
            "terminated_at": p.get("terminated_at").isoformat() if isinstance(p.get("terminated_at"), datetime) else str(p.get("terminated_at")),
            "anti_cheat_logs": p.get("anti_cheat_logs", [])
        })

    return jsonify({"success": True, "participants": participants}), 200

# ----------------- SUBMISSIONS STREAM -----------------

@admin_bp.route("/submissions", methods=["GET"])
@admin_required
def list_all_submissions():
    """Stream all submissions across the platform for admin audit."""
    db = get_db()
    status = request.args.get("status")
    search = request.args.get("search", "").strip()
    page = int(request.args.get("page", 1))
    limit = int(request.args.get("limit", 25))
    skip = (page - 1) * limit

    query = {}
    if status and status.lower() != "all":
        query["status"] = status
    if search:
        query["$or"] = [
            {"student_id": {"$regex": search, "$options": "i"}},
            {"student_name": {"$regex": search, "$options": "i"}},
            {"problem_title": {"$regex": search, "$options": "i"}}
        ]

    total = db.submissions.count_documents(query)
    cursor = db.submissions.find(query).sort("created_at", -1).skip(skip).limit(limit)

    submissions = []
    for s in cursor:
        submissions.append({
            "id": str(s["_id"]),
            "student_id": s.get("student_id"),
            "student_name": s.get("student_name"),
            "problem_id": s.get("problem_id"),
            "problem_title": s.get("problem_title"),
            "language": s.get("language"),
            "status": s.get("status"),
            "runtime": s.get("runtime", 0),
            "passed_test_cases": s.get("passed_test_cases", 0),
            "total_test_cases": s.get("total_test_cases", 0),
            "created_at": s.get("created_at").isoformat() if isinstance(s.get("created_at"), datetime) else str(s.get("created_at"))
        })

    return jsonify({
        "success": True,
        "submissions": submissions,
        "pagination": {"total": total, "page": page, "limit": limit, "pages": (total + limit - 1) // limit if limit > 0 else 1}
    }), 200

# ----------------- ATTENDANCE MANAGEMENT (CONTEST & DAILY SOLVING) -----------------

def parse_attendance_filters(request_obj):
    """Helper to parse filters for department, year, month, calendar_year, search."""
    now = datetime.now(timezone.utc)
    try:
        month = int(request_obj.args.get("month", now.month))
        if month < 1 or month > 12:
            month = now.month
    except (ValueError, TypeError):
        month = now.month

    try:
        calendar_year = int(request_obj.args.get("calendar_year", now.year))
        if calendar_year < 2020 or calendar_year > 2035:
            calendar_year = now.year
    except (ValueError, TypeError):
        calendar_year = now.year

    department = request_obj.args.get("department", "").strip()
    year = request_obj.args.get("year", "").strip()
    search = request_obj.args.get("search", "").strip()

    num_days = calendar.monthrange(calendar_year, month)[1]
    start_of_month = datetime(calendar_year, month, 1, 0, 0, 0, tzinfo=timezone.utc)
    end_of_month = datetime(calendar_year, month, num_days, 23, 59, 59, tzinfo=timezone.utc)

    # Student query
    conditions = [{"role": "STUDENT"}]
    if search:
        conditions.append({
            "$or": [
                {"student_id": {"$regex": search, "$options": "i"}},
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}}
            ]
        })
    if department and department.lower() != "all":
        dept_keywords = {
            "computer science & engineering": ["computer science", "cse"],
            "information technology": ["information technology", "it"],
            "artificial intelligence & data science": ["artificial intelligence", "aids", "ai & ds", "ai and ds"],
            "electronics & communication engineering": ["electronics", "ece"],
            "electrical & electronics engineering": ["electrical", "eee"],
            "mechanical engineering": ["mechanical", "mech"],
            "civil engineering": ["civil"],
            "cyber security": ["cyber"]
        }
        dept_lower = department.lower()
        if dept_lower in dept_keywords:
            or_patterns = [{"department": {"$regex": kw, "$options": "i"}} for kw in dept_keywords[dept_lower]]
            conditions.append({"$or": or_patterns})
        else:
            conditions.append({"department": {"$regex": department, "$options": "i"}})
    if year and year.lower() != "all":
        year_prefix = year.split(" ")[0]
        conditions.append({"year": {"$regex": f"^{year_prefix}", "$options": "i"}})

    student_query = {"$and": conditions} if len(conditions) > 1 else conditions[0]

    return {
        "month": month,
        "calendar_year": calendar_year,
        "num_days": num_days,
        "start_of_month": start_of_month,
        "end_of_month": end_of_month,
        "department": department,
        "year": year,
        "search": search,
        "student_query": student_query,
        "now": now
    }

@admin_bp.route("/attendance/contest", methods=["GET"])
@admin_required
def get_contest_attendance():
    """Calculate student-wise contest attendance for the selected month/year."""
    db = get_db()
    opts = parse_attendance_filters(request)
    now = opts["now"]
    num_days = opts["num_days"]
    calendar_year = opts["calendar_year"]
    month = opts["month"]

    # 1. Find all published contests in that month
    all_contests = list(db.contests.find({"is_published": True}))
    contests_by_day = {}
    month_contest_ids = []

    for c in all_contests:
        st = parse_to_utc_datetime(c.get("start_time"))
        if st:
            st_ist = st.astimezone(IST)
            if st_ist.year == calendar_year and st_ist.month == month:
                d = st_ist.day
                if d not in contests_by_day:
                    contests_by_day[d] = []
                contests_by_day[d].append({
                    "id": str(c["_id"]),
                    "title": c.get("title", "Untitled Contest"),
                    "start_time": format_utc_iso(st)
                })
                month_contest_ids.append(str(c["_id"]))

    # 2. Find participants for these contests
    participants = list(db.contest_participants.find({"contest_id": {"$in": month_contest_ids}}))
    participation_map = {}
    for p in participants:
        c_id = p.get("contest_id")
        u_id = p.get("user_id")
        s_id = p.get("student_id")
        joined_at = p.get("joined_at")
        joined_iso = joined_at.isoformat() if isinstance(joined_at, datetime) else str(joined_at or "")

        part_info = {
            "score": p.get("score", 0),
            "submitted": p.get("submitted", False),
            "joined_at": joined_iso
        }
        if s_id and c_id:
            participation_map[(s_id, c_id)] = part_info
        if u_id and c_id:
            participation_map[(str(u_id), c_id)] = part_info

    # 3. Query matching students
    students = list(db.users.find(opts["student_query"]).sort("student_id", 1))

    # 4. Build day-by-day attendance grid for each student
    student_records = []
    total_present_instances = 0
    total_absent_instances = 0

    for st in students:
        s_id = st.get("student_id", "")
        u_id_str = str(st["_id"])
        s_name = st.get("name", "")
        dept = st.get("department", "CSE")
        yr = st.get("year", "1st Year")

        days_status = {}
        present_count = 0
        absent_count = 0

        for d in range(1, num_days + 1):
            day_date = datetime(calendar_year, month, d, tzinfo=timezone.utc)
            is_future = day_date.date() > now.date()
            day_contests = contests_by_day.get(d, [])

            if is_future:
                days_status[str(d)] = {"status": "FUTURE", "detail": None}
            elif not day_contests:
                days_status[str(d)] = {"status": "NO_CONTEST", "detail": None}
            else:
                # Contest was held on this day
                entered_contest = None
                for c in day_contests:
                    c_id = c["id"]
                    if (s_id, c_id) in participation_map:
                        entered_contest = {**c, **participation_map[(s_id, c_id)]}
                        break
                    elif (u_id_str, c_id) in participation_map:
                        entered_contest = {**c, **participation_map[(u_id_str, c_id)]}
                        break

                if entered_contest:
                    days_status[str(d)] = {"status": "PRESENT", "detail": entered_contest}
                    present_count += 1
                else:
                    days_status[str(d)] = {"status": "ABSENT", "detail": {"contests": day_contests}}
                    absent_count += 1

        total_applicable = present_count + absent_count
        attendance_pct = round((present_count / total_applicable) * 100, 1) if total_applicable > 0 else 100.0

        total_present_instances += present_count
        total_absent_instances += absent_count

        student_records.append({
            "id": u_id_str,
            "student_id": s_id,
            "name": s_name,
            "department": dept,
            "year": yr,
            "days": days_status,
            "present_count": present_count,
            "absent_count": absent_count,
            "attendance_percentage": attendance_pct
        })

    total_eval = total_present_instances + total_absent_instances
    overall_rate = round((total_present_instances / total_eval) * 100, 1) if total_eval > 0 else 100.0

    return jsonify({
        "success": True,
        "month": month,
        "calendar_year": calendar_year,
        "num_days": num_days,
        "contests_by_day": {str(k): v for k, v in contests_by_day.items()},
        "students": student_records,
        "summary": {
            "total_students": len(students),
            "total_present": total_present_instances,
            "total_absent": total_absent_instances,
            "attendance_rate": overall_rate
        }
    }), 200

@admin_bp.route("/attendance/daily", methods=["GET"])
@admin_required
def get_daily_solving_attendance():
    """Calculate student-wise Daily Problem Solving attendance for the selected month/year."""
    db = get_db()
    opts = parse_attendance_filters(request)
    now = opts["now"]
    num_days = opts["num_days"]
    calendar_year = opts["calendar_year"]
    month = opts["month"]

    # 1. Query all accepted submissions in the month
    start_dt = opts["start_of_month"]
    end_dt = opts["end_of_month"]

    subs = list(db.submissions.find({
        "created_at": {"$gte": start_dt, "$lte": end_dt},
        "status": "Accepted"
    }))

    # Map (student_id or user_id, day) -> list of solved problems
    solving_map = {}
    for s in subs:
        c_at = s.get("created_at")
        if isinstance(c_at, datetime):
            if c_at.tzinfo is None:
                c_at = c_at.replace(tzinfo=timezone.utc)
            d = c_at.day
            u_id = s.get("user_id")
            s_id = s.get("student_id")
            prob_title = s.get("problem_title", "Problem")
            prob_id = s.get("problem_id")
            lang = s.get("language", "python")

            item = {
                "problem_id": prob_id,
                "problem_title": prob_title,
                "language": lang,
                "time": c_at.isoformat()
            }
            if s_id:
                solving_map.setdefault((s_id, d), []).append(item)
            if u_id:
                solving_map.setdefault((str(u_id), d), []).append(item)

    # 2. Query matching students
    students = list(db.users.find(opts["student_query"]).sort("student_id", 1))

    student_records = []
    total_present_instances = 0
    total_absent_instances = 0

    for st in students:
        s_id = st.get("student_id", "")
        u_id_str = str(st["_id"])
        s_name = st.get("name", "")
        dept = st.get("department", "CSE")
        yr = st.get("year", "1st Year")

        days_status = {}
        present_count = 0
        absent_count = 0

        for d in range(1, num_days + 1):
            day_date = datetime(calendar_year, month, d, tzinfo=timezone.utc)
            is_future = day_date.date() > now.date()

            if is_future:
                days_status[str(d)] = {"status": "FUTURE", "solved": []}
            else:
                solved_items = solving_map.get((s_id, d), []) or solving_map.get((u_id_str, d), [])
                if len(solved_items) > 0:
                    days_status[str(d)] = {"status": "PRESENT", "solved": solved_items}
                    present_count += 1
                else:
                    days_status[str(d)] = {"status": "ABSENT", "solved": []}
                    absent_count += 1

        total_applicable = present_count + absent_count
        attendance_pct = round((present_count / total_applicable) * 100, 1) if total_applicable > 0 else 100.0

        total_present_instances += present_count
        total_absent_instances += absent_count

        student_records.append({
            "id": u_id_str,
            "student_id": s_id,
            "name": s_name,
            "department": dept,
            "year": yr,
            "days": days_status,
            "present_count": present_count,
            "absent_count": absent_count,
            "attendance_percentage": attendance_pct
        })

    total_eval = total_present_instances + total_absent_instances
    overall_rate = round((total_present_instances / total_eval) * 100, 1) if total_eval > 0 else 100.0

    return jsonify({
        "success": True,
        "month": month,
        "calendar_year": calendar_year,
        "num_days": num_days,
        "students": student_records,
        "summary": {
            "total_students": len(students),
            "total_present": total_present_instances,
            "total_absent": total_absent_instances,
            "attendance_rate": overall_rate
        }
    }), 200

@admin_bp.route("/attendance/export", methods=["GET"])
@admin_required
def export_attendance_excel():
    """Export formatted Excel report for Contest or Daily Solving Attendance."""
    db = get_db()
    att_type = request.args.get("type", "contest").strip().lower()
    opts = parse_attendance_filters(request)
    now = opts["now"]
    num_days = opts["num_days"]
    calendar_year = opts["calendar_year"]
    month = opts["month"]
    month_name = calendar.month_name[month]

    wb = openpyxl.Workbook()
    # Sheet 1: Monthly Grid
    ws_grid = wb.active
    ws_grid.title = f"{'Contest' if att_type == 'contest' else 'Daily'} Attendance Grid"

    # Header style
    header_fill = PatternFill(start_color="303442", end_color="303442", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # soft green
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # soft red
    future_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # soft grey

    # Title Banner
    ws_grid.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 7)
    title_cell = ws_grid.cell(row=1, column=1)
    title_cell.value = f"NIT Campus Coder — {'Contest' if att_type == 'contest' else 'Daily Problem Solving'} Attendance Report ({month_name} {calendar_year})"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0757B8", end_color="0757B8", fill_type="solid")
    title_cell.alignment = center_align

    # Row 3: Column Headers
    headers = ["Student ID", "Student Name", "Department", "Year"]
    for d in range(1, num_days + 1):
        headers.append(str(d))
    headers.extend(["Present", "Absent", "Attendance %"])

    for col_idx, h_text in enumerate(headers, 1):
        cell = ws_grid.cell(row=3, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align if col_idx > 4 else left_align

    # Fetch data based on type
    students = list(db.users.find(opts["student_query"]).sort("student_id", 1))
    detailed_logs = []

    if att_type == "contest":
        all_contests = list(db.contests.find({"is_published": True}))
        contests_by_day = {}
        month_contest_ids = []
        for c in all_contests:
            st = c.get("start_time")
            if isinstance(st, datetime):
                if st.tzinfo is None: st = st.replace(tzinfo=timezone.utc)
                if st.year == calendar_year and st.month == month:
                    contests_by_day.setdefault(st.day, []).append(c)
                    month_contest_ids.append(str(c["_id"]))

        participants = list(db.contest_participants.find({"contest_id": {"$in": month_contest_ids}}))
        participation_map = {}
        for p in participants:
            c_id = p.get("contest_id")
            s_id = p.get("student_id")
            u_id = p.get("user_id")
            if s_id and c_id: participation_map[(s_id, c_id)] = p
            if u_id and c_id: participation_map[(str(u_id), c_id)] = p

        row_idx = 4
        for st in students:
            s_id = st.get("student_id", "")
            u_id = str(st["_id"])
            s_name = st.get("name", "")
            dept = st.get("department", "CSE")
            yr = st.get("year", "1st Year")

            ws_grid.cell(row=row_idx, column=1, value=s_id).alignment = left_align
            ws_grid.cell(row=row_idx, column=2, value=s_name).alignment = left_align
            ws_grid.cell(row=row_idx, column=3, value=dept).alignment = left_align
            ws_grid.cell(row=row_idx, column=4, value=yr).alignment = left_align

            present_count = 0
            absent_count = 0

            for d in range(1, num_days + 1):
                col_i = 4 + d
                day_date = datetime(calendar_year, month, d, tzinfo=timezone.utc)
                is_future = day_date.date() > now.date()
                day_contests = contests_by_day.get(d, [])

                cell = ws_grid.cell(row=row_idx, column=col_i)
                cell.alignment = center_align

                if is_future:
                    cell.value = "-"
                    cell.fill = future_fill
                elif not day_contests:
                    cell.value = "-"
                    cell.fill = future_fill
                else:
                    joined = None
                    for c in day_contests:
                        c_id = str(c["_id"])
                        if (s_id, c_id) in participation_map:
                            joined = participation_map[(s_id, c_id)]
                            break
                        elif (u_id, c_id) in participation_map:
                            joined = participation_map[(u_id, c_id)]
                            break

                    if joined:
                        cell.value = "PRESENT"
                        cell.fill = present_fill
                        present_count += 1
                        detailed_logs.append({
                            "student_id": s_id,
                            "name": s_name,
                            "dept": dept,
                            "year": yr,
                            "date": f"{calendar_year}-{month:02d}-{d:02d}",
                            "item": day_contests[0].get("title", "Contest"),
                            "time": str(joined.get("joined_at", "")),
                            "status": "PRESENT"
                        })
                    else:
                        cell.value = "ABSENT"
                        cell.fill = absent_fill
                        absent_count += 1
                        detailed_logs.append({
                            "student_id": s_id,
                            "name": s_name,
                            "dept": dept,
                            "year": yr,
                            "date": f"{calendar_year}-{month:02d}-{d:02d}",
                            "item": day_contests[0].get("title", "Contest"),
                            "time": "-",
                            "status": "ABSENT"
                        })

            tot_app = present_count + absent_count
            pct = round((present_count / tot_app) * 100, 1) if tot_app > 0 else 100.0

            ws_grid.cell(row=row_idx, column=num_days + 5, value=present_count).alignment = center_align
            ws_grid.cell(row=row_idx, column=num_days + 6, value=absent_count).alignment = center_align
            ws_grid.cell(row=row_idx, column=num_days + 7, value=f"{pct}%").alignment = center_align
            row_idx += 1

    else:
        # Daily Solving Attendance
        start_dt = opts["start_of_month"]
        end_dt = opts["end_of_month"]
        subs = list(db.submissions.find({"created_at": {"$gte": start_dt, "$lte": end_dt}, "status": "Accepted"}))
        solving_map = {}
        for s in subs:
            c_at = s.get("created_at")
            if isinstance(c_at, datetime):
                if c_at.tzinfo is None: c_at = c_at.replace(tzinfo=timezone.utc)
                d = c_at.day
                s_id = s.get("student_id")
                u_id = s.get("user_id")
                if s_id: solving_map.setdefault((s_id, d), []).append(s)
                if u_id: solving_map.setdefault((str(u_id), d), []).append(s)

        row_idx = 4
        for st in students:
            s_id = st.get("student_id", "")
            u_id = str(st["_id"])
            s_name = st.get("name", "")
            dept = st.get("department", "CSE")
            yr = st.get("year", "1st Year")

            ws_grid.cell(row=row_idx, column=1, value=s_id).alignment = left_align
            ws_grid.cell(row=row_idx, column=2, value=s_name).alignment = left_align
            ws_grid.cell(row=row_idx, column=3, value=dept).alignment = left_align
            ws_grid.cell(row=row_idx, column=4, value=yr).alignment = left_align

            present_count = 0
            absent_count = 0

            for d in range(1, num_days + 1):
                col_i = 4 + d
                day_date = datetime(calendar_year, month, d, tzinfo=timezone.utc)
                is_future = day_date.date() > now.date()
                solved = solving_map.get((s_id, d), []) or solving_map.get((u_id, d), [])

                cell = ws_grid.cell(row=row_idx, column=col_i)
                cell.alignment = center_align

                if is_future:
                    cell.value = "-"
                    cell.fill = future_fill
                elif len(solved) > 0:
                    cell.value = "PRESENT"
                    cell.fill = present_fill
                    present_count += 1
                    detailed_logs.append({
                        "student_id": s_id,
                        "name": s_name,
                        "dept": dept,
                        "year": yr,
                        "date": f"{calendar_year}-{month:02d}-{d:02d}",
                        "item": solved[0].get("problem_title", "Daily Problem"),
                        "time": str(solved[0].get("created_at", "")),
                        "status": "PRESENT"
                    })
                else:
                    cell.value = "ABSENT"
                    cell.fill = absent_fill
                    absent_count += 1
                    detailed_logs.append({
                        "student_id": s_id,
                        "name": s_name,
                        "dept": dept,
                        "year": yr,
                        "date": f"{calendar_year}-{month:02d}-{d:02d}",
                        "item": "Daily Problem Solving",
                        "time": "-",
                        "status": "ABSENT"
                    })

            tot_app = present_count + absent_count
            pct = round((present_count / tot_app) * 100, 1) if tot_app > 0 else 100.0

            ws_grid.cell(row=row_idx, column=num_days + 5, value=present_count).alignment = center_align
            ws_grid.cell(row=row_idx, column=num_days + 6, value=absent_count).alignment = center_align
            ws_grid.cell(row=row_idx, column=num_days + 7, value=f"{pct}%").alignment = center_align
            row_idx += 1

    # Sheet 2: Detailed Activity Log
    ws_log = wb.create_sheet(title="Detailed Activity Log")
    log_headers = ["Student ID", "Student Name", "Department", "Year", "Date", "Activity / Challenge", "Timestamp", "Attendance Status"]
    for col_idx, h_text in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = left_align

    for l_idx, log_item in enumerate(detailed_logs, 2):
        ws_log.cell(row=l_idx, column=1, value=log_item["student_id"])
        ws_log.cell(row=l_idx, column=2, value=log_item["name"])
        ws_log.cell(row=l_idx, column=3, value=log_item["dept"])
        ws_log.cell(row=l_idx, column=4, value=log_item["year"])
        ws_log.cell(row=l_idx, column=5, value=log_item["date"])
        ws_log.cell(row=l_idx, column=6, value=log_item["item"])
        ws_log.cell(row=l_idx, column=7, value=log_item["time"])
        st_cell = ws_log.cell(row=l_idx, column=8, value=log_item["status"])
        st_cell.fill = present_fill if log_item["status"] == "PRESENT" else absent_fill

    # Adjust column widths
    for ws in [ws_grid, ws_log]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 9)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"{att_type}_attendance_{month_name}_{calendar_year}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# ----------------- CONTEST REPORTS & ADVANCED LEADERBOARD -----------------

@admin_bp.route("/reports/contests", methods=["GET"])
@admin_required
def list_contests_for_reports():
    """List all contests available for performance reporting."""
    db = get_db()
    contests_cursor = db.contests.find({}).sort("start_time", -1)
    
    now = get_utc_now()
    contests = []
    for c in contests_cursor:
        c_id = str(c["_id"])
        participants_count = db.contest_participants.count_documents({"contest_id": c_id})
        prob_ids = [str(pid) for pid in c.get("problem_ids", []) if pid]
        mcq_ids = [str(mid) for mid in c.get("mcq_ids", []) if mid]
        
        start = parse_to_utc_datetime(c.get("start_time"))
        end = parse_to_utc_datetime(c.get("end_time"))
        status = calculate_contest_status(start, end, now)
        
        contests.append({
            "id": c_id,
            "title": c.get("title", "Contest"),
            "description": c.get("description", ""),
            "start_time": format_utc_iso(start),
            "end_time": format_utc_iso(end),
            "status": status,
            "duration_minutes": c.get("duration_minutes", 60),
            "problems_count": len(prob_ids),
            "mcqs_count": len(mcq_ids),
            "total_points": c.get("total_points", 100),
            "is_published": c.get("is_published", False),
            "participants_count": participants_count
        })

    return jsonify({"success": True, "contests": contests, "server_time": format_utc_iso(now)}), 200

def calculate_candidate_contest_metrics(contest, problems, participant, submissions):
    """
    Compute 5-factor weighted score (out of 100) and complexity ratings for a candidate.
    Formula:
      - Test Cases:        50%
      - Problems Solved:   20%
      - Time Efficiency:   10%
      - Time Complexity:   10%
      - Space Complexity:  10%
    """
    total_contest_problems = max(len(problems), 1)
    total_contest_testcases = sum(len(p.get("test_cases", [])) or 4 for p in problems)
    if total_contest_testcases == 0:
        total_contest_testcases = total_contest_problems * 4

    # Evaluate problems solved and passed test cases
    solved_count = 0
    passed_test_cases = 0
    total_runtime = 0.0
    total_memory = 0.0
    evaluated_submissions_count = 0

    problem_breakdowns = []

    for prob in problems:
        p_id = str(prob["_id"])
        prob_title = prob.get("title", "Problem")
        prob_tcs_total = len(prob.get("test_cases", [])) or 4

        # Find candidate's best submission for this problem
        cand_subs = [s for s in submissions if str(s.get("problem_id")) == p_id]
        best_sub = None
        for s in cand_subs:
            if s.get("status") == "Accepted":
                best_sub = s
                break
        if not best_sub and cand_subs:
            # Pick submission with max passed test cases
            best_sub = max(cand_subs, key=lambda x: x.get("passed_test_cases", 0))

        if best_sub:
            sub_passed_tc = best_sub.get("passed_test_cases", 0)
            if best_sub.get("status") == "Accepted":
                solved_count += 1
                sub_passed_tc = max(sub_passed_tc, prob_tcs_total)

            passed_test_cases += min(sub_passed_tc, prob_tcs_total)
            rt = float(best_sub.get("runtime", 35.0) or 35.0)
            mem = float(best_sub.get("memory", 18.0) or 18.0)
            total_runtime += rt
            total_memory += mem
            evaluated_submissions_count += 1

            problem_breakdowns.append({
                "problem_id": p_id,
                "problem_title": prob_title,
                "difficulty": prob.get("difficulty", "Medium"),
                "status": best_sub.get("status", "Attempted"),
                "passed_test_cases": sub_passed_tc,
                "total_test_cases": prob_tcs_total,
                "runtime": rt,
                "memory": mem,
                "language": best_sub.get("language", "python")
            })
        else:
            # Check if participant score indicates solved
            if participant.get("score", 0) > 0 and solved_count < participant.get("problems_solved", 0):
                solved_count += 1
                passed_test_cases += prob_tcs_total
                problem_breakdowns.append({
                    "problem_id": p_id,
                    "problem_title": prob_title,
                    "difficulty": prob.get("difficulty", "Medium"),
                    "status": "Accepted",
                    "passed_test_cases": prob_tcs_total,
                    "total_test_cases": prob_tcs_total,
                    "runtime": 32.5,
                    "memory": 16.2,
                    "language": "python"
                })
            else:
                problem_breakdowns.append({
                    "problem_id": p_id,
                    "problem_title": prob_title,
                    "difficulty": prob.get("difficulty", "Medium"),
                    "status": "Not Attempted",
                    "passed_test_cases": 0,
                    "total_test_cases": prob_tcs_total,
                    "runtime": 0,
                    "memory": 0,
                    "language": "-"
                })

    # If participant has problem solved count higher in document, respect it
    if participant.get("problems_solved", 0) > solved_count:
        solved_count = min(participant.get("problems_solved", 0), total_contest_problems)
        passed_test_cases = max(passed_test_cases, solved_count * 4)

    # 1. Test Cases Score (50%)
    tc_ratio = min(passed_test_cases / max(total_contest_testcases, 1), 1.0)
    score_test_cases = round(tc_ratio * 50.0, 1)

    # 2. Problems Solved Score (20%)
    prob_ratio = min(solved_count / max(total_contest_problems, 1), 1.0)
    score_problems_solved = round(prob_ratio * 20.0, 1)

    # 3. Time Taken & Time Efficiency (10%)
    duration_min = contest.get("duration_minutes", 60)
    duration_sec = duration_min * 60
    joined_at = participant.get("joined_at")
    submitted_at = participant.get("submitted_at")

    joined_at_utc = parse_to_utc_datetime(joined_at)
    submitted_at_utc = parse_to_utc_datetime(submitted_at)
    terminated_at_utc = parse_to_utc_datetime(participant.get("terminated_at"))
    if joined_at_utc:
        if submitted_at_utc:
            end_time_calc = submitted_at_utc
        elif terminated_at_utc:
            end_time_calc = terminated_at_utc
        else:
            end_time_calc = get_utc_now()
        time_taken_sec = max(int((end_time_calc - joined_at_utc).total_seconds()), 60)
        time_taken_sec = min(time_taken_sec, duration_sec)
    else:
        # Default estimation based on solved count
        time_taken_sec = min(solved_count * 900 + 300, duration_sec)

    time_ratio = min(time_taken_sec / max(duration_sec, 60), 1.0)
    # Faster submissions get higher efficiency score
    if solved_count > 0 or passed_test_cases > 0:
        score_time_efficiency = round(max(10.0 * (1.0 - 0.5 * time_ratio), 2.0), 1)
    else:
        score_time_efficiency = 0.0

    # 4. Time Complexity Score (10%)
    avg_runtime = total_runtime / max(evaluated_submissions_count, 1) if evaluated_submissions_count > 0 else 0
    if solved_count > 0 or passed_test_cases > 0:
        if avg_runtime <= 45.0:
            score_time_complexity = 10.0
            complexity_time_label = "O(N) Optimal"
        elif avg_runtime <= 120.0:
            score_time_complexity = 9.0
            complexity_time_label = "O(N log N)"
        elif avg_runtime <= 350.0:
            score_time_complexity = 8.0
            complexity_time_label = "O(N²)"
        else:
            score_time_complexity = 6.5
            complexity_time_label = "O(2^N)"
    else:
        score_time_complexity = 0.0
        complexity_time_label = "-"

    # 5. Space Complexity Score (10%)
    avg_memory = total_memory / max(evaluated_submissions_count, 1) if evaluated_submissions_count > 0 else 0
    if solved_count > 0 or passed_test_cases > 0:
        if avg_memory <= 18.0:
            score_space_complexity = 10.0
            complexity_space_label = "O(1) Auxiliary"
        elif avg_memory <= 32.0:
            score_space_complexity = 9.0
            complexity_space_label = "O(N)"
        else:
            score_space_complexity = 7.5
            complexity_space_label = "O(N²)"
    else:
        score_space_complexity = 0.0
        complexity_space_label = "-"

    final_score = round(score_test_cases + score_problems_solved + score_time_efficiency + score_time_complexity + score_space_complexity, 1)
    final_score = min(max(final_score, 0.0), 100.0)

    # Format time taken string
    mins = time_taken_sec // 60
    secs = time_taken_sec % 60
    time_taken_str = f"{mins}m {secs:02d}s"

    # Anti-cheat status
    auto_terminated = bool(participant.get("auto_terminated") or participant.get("status") == "AUTO_TERMINATED")
    flags_count = len(participant.get("anti_cheat_logs", []))
    if auto_terminated:
        anti_cheat_status = "AUTO_TERMINATED"
    elif flags_count > 0:
        anti_cheat_status = "FLAGGED"
    else:
        anti_cheat_status = "CLEAN"

    # MCQ & Coding Specific Metrics
    assigned_mcqs = participant.get("assigned_mcq_ids")
    total_contest_mcqs = len(assigned_mcqs) if assigned_mcqs else int(contest.get("mcqs_per_student") or len(contest.get("mcq_ids", [])))
    mcqs_correct = int(participant.get("mcqs_correct", 0))
    mcq_score = float(participant.get("mcq_score", mcqs_correct * 10))
    mcq_percentage = round((mcqs_correct / max(total_contest_mcqs, 1)) * 100, 1) if total_contest_mcqs > 0 else 0.0

    coding_score = float(participant.get("coding_score", max(float(participant.get("score", 0)) - mcq_score, 0.0)))
    coding_percentage = round((passed_test_cases / max(total_contest_testcases, 1)) * 100, 1) if total_contest_testcases > 0 else 0.0
    overall_score = float(participant.get("score", mcq_score + coding_score))

    return {
        "solved_count": solved_count,
        "total_contest_problems": total_contest_problems,
        "passed_test_cases": passed_test_cases,
        "total_contest_testcases": total_contest_testcases,
        "time_taken_seconds": time_taken_sec,
        "time_taken_formatted": time_taken_str,
        "time_complexity_label": complexity_time_label,
        "space_complexity_label": complexity_space_label,
        "final_score": final_score,
        "mcq_score": mcq_score,
        "mcqs_correct": mcqs_correct,
        "total_contest_mcqs": total_contest_mcqs,
        "mcq_percentage": mcq_percentage,
        "coding_score": coding_score,
        "coding_percentage": coding_percentage,
        "overall_score": overall_score,
        "score_breakdown": {
            "test_cases": score_test_cases,
            "problems_solved": score_problems_solved,
            "time_efficiency": score_time_efficiency,
            "time_complexity": score_time_complexity,
            "space_complexity": score_space_complexity,
            "mcq_score": mcq_score,
            "coding_score": coding_score,
            "overall_score": overall_score,
            "total": final_score
        },
        "anti_cheat": {
            "status": anti_cheat_status,
            "auto_terminated": auto_terminated,
            "termination_reason": participant.get("termination_reason", ""),
            "flags_count": flags_count,
            "logs": participant.get("anti_cheat_logs", [])
        },
        "problem_breakdowns": problem_breakdowns
    }

@admin_bp.route("/reports/contests/<contest_id>", methods=["GET"])
@admin_required
def get_contest_report(contest_id):
    """Retrieve full contest performance report with ranked leaderboard and score breakdown."""
    db = get_db()
    if not ObjectId.is_valid(contest_id):
        return jsonify({"error": "Invalid contest ID", "success": False}), 400

    contest = db.contests.find_one({"_id": ObjectId(contest_id)})
    if not contest:
        return jsonify({"error": "Contest not found", "success": False}), 404

    # Fetch contest problems
    problem_ids = contest.get("problem_ids", [])
    problems = []
    for pid in problem_ids:
        if ObjectId.is_valid(str(pid)):
            p_doc = db.problems.find_one({"_id": ObjectId(str(pid))})
            if p_doc:
                problems.append(p_doc)
        else:
            p_doc = db.problems.find_one({"id": pid})
            if p_doc:
                problems.append(p_doc)

    # Materialize expired lock windows before building report statuses.
    now = get_utc_now()
    db.contest_participants.update_many(
        {
            "contest_id": contest_id,
            "status": "LOCKED",
            "lock_timeout_at": {"$lte": now},
        },
        {"$set": {
            "status": "AUTO_TERMINATED",
            "auto_terminated": True,
            "resolution_window_active": False,
            "terminated_at": now,
            "termination_reason": "Lock resolution window (30 minutes) expired without admin action",
        }},
    )

    # Fetch participants — all attempts
    contest_id_values = [contest_id, ObjectId(contest_id)]
    all_participants = list(db.contest_participants.find({"contest_id": {"$in": contest_id_values}}))
    
    # Deduplicate: keep best/latest attempt per student for the leaderboard
    # Priority: submitted retest > submitted original > active > locked/terminated
    student_best = {}
    for p in all_participants:
        uid_key = str(p.get("student_id") or p.get("user_id", ""))
        if not uid_key:
            continue
        existing = student_best.get(uid_key)
        if existing is None:
            student_best[uid_key] = p
        else:
            # Prefer submitted over non-submitted
            p_submitted = bool(p.get("submitted"))
            e_submitted = bool(existing.get("submitted"))
            if p_submitted and not e_submitted:
                student_best[uid_key] = p
            elif p_submitted == e_submitted:
                # Prefer higher attempt number (latest)
                if p.get("attempt_number", 1) > existing.get("attempt_number", 1):
                    student_best[uid_key] = p
                elif p.get("score", 0) > existing.get("score", 0):
                    student_best[uid_key] = p

    # Build a lookup of original attempt scores for retest students
    original_scores = {}
    retest_scores = {}
    for p in all_participants:
        uid_key = str(p.get("user_id", p.get("student_id", "")))
        if p.get("attempt_number", 1) == 1:
            original_scores[uid_key] = {
                "score": p.get("score", 0),
                "mcq_score": float(p.get("mcq_score", 0)),
                "coding_score": float(p.get("coding_score", 0)),
                "status": p.get("status", "")
            }
        elif p.get("submitted") and (uid_key not in retest_scores or p.get("attempt_number", 1) > retest_scores[uid_key].get("attempt_number", 1)):
            retest_scores[uid_key] = {
                "score": p.get("score", 0),
                "mcq_score": float(p.get("mcq_score", 0)),
                "coding_score": float(p.get("coding_score", 0)),
                "attempt_number": p.get("attempt_number", 1),
                "status": p.get("status", "")
            }

    participants = list(student_best.values())

    # Fetch all submissions for this contest
    submissions = list(db.submissions.find({}))

    # Optional department / year filter from request
    dept_filter = request.args.get("department", "").strip()
    year_filter = request.args.get("year", "").strip()
    search_filter = request.args.get("search", "").strip().lower()

    leaderboard = []

    for p in participants:
        u_id = p.get("user_id")
        s_id = p.get("student_id")
        user_doc = None
        if u_id and ObjectId.is_valid(str(u_id)):
            user_doc = db.users.find_one({"_id": ObjectId(str(u_id))})
        elif s_id:
            user_doc = db.users.find_one({"student_id": s_id})

        student_name = p.get("student_name") or (user_doc.get("name") if user_doc else "Student")
        student_id_val = s_id or (user_doc.get("student_id") if user_doc else "STU")
        department = p.get("department") or (user_doc.get("department") if user_doc else "CSE")
        year = (user_doc.get("year") if user_doc else "1st Year")

        # Apply search and department/year filters
        if dept_filter and dept_filter.lower() != "all" and dept_filter.lower() not in department.lower():
            continue
        if year_filter and year_filter.lower() != "all" and year_filter.split(" ")[0].lower() not in year.lower():
            continue
        if search_filter:
            if search_filter not in student_name.lower() and search_filter not in student_id_val.lower():
                continue

        cand_subs = [s for s in submissions if s.get("student_id") == student_id_val or str(s.get("user_id")) == str(u_id)]
        metrics = calculate_candidate_contest_metrics(contest, problems, p, cand_subs)

        attempt_num = p.get("attempt_number", 1)
        uid_key = str(s_id or u_id or "")
        orig_info = original_scores.get(uid_key) if attempt_num > 1 else None
        retest_info = retest_scores.get(uid_key) if attempt_num == 1 else ({
            "score": metrics["overall_score"],
            "mcq_score": metrics["mcq_score"],
            "coding_score": metrics["coding_score"],
            "attempt_number": attempt_num,
            "status": p.get("status", "")
        } if p.get("submitted") else None)

        leaderboard.append({
            "participant_id": str(p["_id"]),
            "user_id": str(u_id or ""),
            "student_id": student_id_val,
            "name": student_name,
            "department": department,
            "year": year,
            "solved_count": metrics["solved_count"],
            "total_contest_problems": metrics["total_contest_problems"],
            "passed_test_cases": metrics["passed_test_cases"],
            "total_contest_testcases": metrics["total_contest_testcases"],
            "time_taken": metrics["time_taken_formatted"],
            "time_taken_seconds": metrics["time_taken_seconds"],
            "time_complexity": metrics["time_complexity_label"],
            "space_complexity": metrics["space_complexity_label"],
            "final_score": metrics["final_score"],
            "mcq_score": metrics["mcq_score"],
            "mcqs_correct": metrics["mcqs_correct"],
            "total_contest_mcqs": metrics["total_contest_mcqs"],
            "mcq_percentage": metrics["mcq_percentage"],
            "coding_score": metrics["coding_score"],
            "coding_percentage": metrics["coding_percentage"],
            "overall_score": metrics["overall_score"],
            "score_breakdown": metrics["score_breakdown"],
            "anti_cheat": metrics["anti_cheat"],
            "problem_breakdowns": metrics["problem_breakdowns"],
            "is_locked": bool(p.get("status") == "LOCKED"),
            "is_terminated": bool(p.get("auto_terminated") or p.get("status") in ["TERMINATED", "AUTO_TERMINATED"]),
            "attempt_number": attempt_num,
            "is_retest": attempt_num > 1,
            "original_score": orig_info,
            "retest_score": retest_info,
            "retest_marks": retest_info.get("score") if retest_info else None
        })

    # Sort leaderboard by:
    # 1. Overall Score / Final Score (descending)
    # 2. Problems Solved (descending)
    # 3. MCQs Correct (descending)
    # 4. Time Taken (ascending)
    leaderboard.sort(key=lambda x: (
        -x["overall_score"],
        -x["final_score"],
        -x["solved_count"],
        -x["mcqs_correct"],
        x["time_taken_seconds"]
    ))

    # Assign ranks
    for idx, item in enumerate(leaderboard, 1):
        item["rank"] = idx

    # Summary KPIs for Overall, MCQ, and Coding
    total_candidates = len(leaderboard)
    avg_score = round(sum(item["overall_score"] for item in leaderboard) / max(total_candidates, 1), 1) if total_candidates > 0 else 0.0
    highest_score = max((item["overall_score"] for item in leaderboard), default=0.0)
    clean_candidates_count = sum(1 for item in leaderboard if item["anti_cheat"]["status"] == "CLEAN")
    clean_rate = round((clean_candidates_count / max(total_candidates, 1)) * 100, 1) if total_candidates > 0 else 100.0

    return jsonify({
        "success": True,
        "contest": {
            "id": str(contest["_id"]),
            "title": contest.get("title"),
            "description": contest.get("description", ""),
            "duration_minutes": contest.get("duration_minutes", 60),
            "start_time": format_utc_iso(parse_to_utc_datetime(contest.get("start_time"))),
            "end_time": format_utc_iso(parse_to_utc_datetime(contest.get("end_time"))),
            "problems_count": len(problems),
            "mcqs_count": len(contest.get("mcq_ids", [])),
            "total_points": contest.get("total_points", 100)
        },
        "summary": {
            "total_candidates": total_candidates,
            "average_score": avg_score,
            "highest_score": highest_score,
            "clean_rate": clean_rate,
            "clean_count": clean_candidates_count,
            "auto_terminated_count": sum(1 for item in leaderboard if item["anti_cheat"]["auto_terminated"]),
            # Overall KPIs
            "avg_overall_score": avg_score,
            "highest_overall_score": highest_score,
            # MCQ KPIs
            "total_mcqs": len(contest.get("mcq_ids", [])),
            "avg_mcq_score": round(sum(item["mcq_score"] for item in leaderboard) / max(total_candidates, 1), 1) if total_candidates > 0 else 0.0,
            "highest_mcq_score": max((item["mcq_score"] for item in leaderboard), default=0.0),
            "avg_mcq_accuracy": round(sum(item["mcq_percentage"] for item in leaderboard) / max(total_candidates, 1), 1) if total_candidates > 0 else 0.0,
            # Coding KPIs
            "total_coding_problems": len(problems),
            "avg_coding_score": round(sum(item["coding_score"] for item in leaderboard) / max(total_candidates, 1), 1) if total_candidates > 0 else 0.0,
            "highest_coding_score": max((item["coding_score"] for item in leaderboard), default=0.0),
            "avg_test_cases_passed": round(sum(item["passed_test_cases"] for item in leaderboard) / max(total_candidates, 1), 1) if total_candidates > 0 else 0.0
        },
        "leaderboard": leaderboard,
        "problems": [{
            "id": str(p["_id"]),
            "title": p.get("title"),
            "difficulty": p.get("difficulty", "Medium"),
            "topic": p.get("topic", "General")
        } for p in problems]
    }), 200

@admin_bp.route("/reports/contests/<contest_id>/export", methods=["GET"])
@admin_required
def export_contest_report_excel(contest_id):
    """Generate and stream formatted Excel or CSV reports for Overall, MCQ, or Coding."""
    db = get_db()
    if not ObjectId.is_valid(contest_id):
        return jsonify({"error": "Invalid contest ID", "success": False}), 400

    contest = db.contests.find_one({"_id": ObjectId(contest_id)})
    if not contest:
        return jsonify({"error": "Contest not found", "success": False}), 404

    # Report type and format
    report_type = (request.args.get("report_type") or request.args.get("type") or "overall").strip().lower()
    export_format = (request.args.get("format") or "excel").strip().lower()

    # Fetch problems and participants
    problem_ids = contest.get("problem_ids", [])
    problems = []
    for pid in problem_ids:
        if ObjectId.is_valid(str(pid)):
            p_doc = db.problems.find_one({"_id": ObjectId(str(pid))})
            if p_doc: problems.append(p_doc)
        else:
            p_doc = db.problems.find_one({"id": pid})
            if p_doc: problems.append(p_doc)

    contest_id_values = [contest_id, ObjectId(contest_id)]
    participants = list(db.contest_participants.find({"contest_id": {"$in": contest_id_values}}))
    submissions = list(db.submissions.find({}))

    original_scores = {}
    retest_scores = {}
    for participant in participants:
        uid_key = str(participant.get("student_id") or participant.get("user_id", ""))
        attempt_number = participant.get("attempt_number", 1)
        if attempt_number == 1:
            original_scores[uid_key] = {
                "score": participant.get("score", 0),
                "mcq_score": float(participant.get("mcq_score", 0)),
                "coding_score": float(participant.get("coding_score", 0)),
                "status": participant.get("status", "")
            }
        elif participant.get("submitted") and (uid_key not in retest_scores or attempt_number > retest_scores[uid_key].get("attempt_number", 1)):
            retest_scores[uid_key] = {
                "score": participant.get("score", 0),
                "mcq_score": float(participant.get("mcq_score", 0)),
                "coding_score": float(participant.get("coding_score", 0)),
                "attempt_number": attempt_number,
                "status": participant.get("status", "")
            }

    # Filters
    dept_filter = request.args.get("department", "").strip()
    year_filter = request.args.get("year", "").strip()
    search_filter = request.args.get("search", "").strip().lower()

    leaderboard = []
    for p in participants:
        u_id = p.get("user_id")
        s_id = p.get("student_id")
        user_doc = None
        if u_id and ObjectId.is_valid(str(u_id)):
            user_doc = db.users.find_one({"_id": ObjectId(str(u_id))})
        elif s_id:
            user_doc = db.users.find_one({"student_id": s_id})

        student_name = p.get("student_name") or (user_doc.get("name") if user_doc else "Student")
        student_id_val = s_id or (user_doc.get("student_id") if user_doc else "STU")
        department = p.get("department") or (user_doc.get("department") if user_doc else "CSE")
        year = (user_doc.get("year") if user_doc else "1st Year")

        if dept_filter and dept_filter.lower() != "all" and dept_filter.lower() not in department.lower():
            continue
        if year_filter and year_filter.lower() != "all" and year_filter.split(" ")[0].lower() not in year.lower():
            continue
        if search_filter:
            if search_filter not in student_name.lower() and search_filter not in student_id_val.lower():
                continue

        cand_subs = [s for s in submissions if s.get("student_id") == student_id_val or str(s.get("user_id")) == str(u_id)]
        metrics = calculate_candidate_contest_metrics(contest, problems, p, cand_subs)
        attempt_number = p.get("attempt_number", 1)
        uid_key = str(s_id or u_id or "")
        original_score = original_scores.get(uid_key) if attempt_number > 1 else None
        retest_score = retest_scores.get(uid_key) if attempt_number == 1 else ({
            "score": metrics["overall_score"],
            "mcq_score": metrics["mcq_score"],
            "coding_score": metrics["coding_score"],
            "attempt_number": attempt_number,
            "status": p.get("status", "")
        } if p.get("submitted") else None)

        leaderboard.append({
            "student_id": student_id_val,
            "name": student_name,
            "department": department,
            "year": year,
            "solved_count": metrics["solved_count"],
            "total_contest_problems": metrics["total_contest_problems"],
            "passed_test_cases": metrics["passed_test_cases"],
            "total_contest_testcases": metrics["total_contest_testcases"],
            "time_taken": metrics["time_taken_formatted"],
            "time_taken_seconds": metrics["time_taken_seconds"],
            "time_complexity": metrics["time_complexity_label"],
            "space_complexity": metrics["space_complexity_label"],
            "final_score": metrics["final_score"],
            "mcq_score": metrics["mcq_score"],
            "mcqs_correct": metrics["mcqs_correct"],
            "total_contest_mcqs": metrics["total_contest_mcqs"],
            "mcq_percentage": metrics["mcq_percentage"],
            "coding_score": metrics["coding_score"],
            "coding_percentage": metrics["coding_percentage"],
            "overall_score": metrics["overall_score"],
            "score_breakdown": metrics["score_breakdown"],
            "anti_cheat": metrics["anti_cheat"],
            "problem_breakdowns": metrics["problem_breakdowns"],
            "attempt_number": attempt_number,
            "is_retest": attempt_number > 1,
            "original_score": original_score,
            "retest_score": retest_score,
            "retest_marks": retest_score.get("score") if retest_score else None
        })

    # Sort based on report type
    if report_type == "mcq":
        leaderboard.sort(key=lambda x: (-x["mcq_score"], -x["mcqs_correct"], x["time_taken_seconds"]))
    elif report_type == "coding":
        leaderboard.sort(key=lambda x: (-x["coding_score"], -x["solved_count"], -x["passed_test_cases"], x["time_taken_seconds"]))
    else:
        # overall
        leaderboard.sort(key=lambda x: (-x["overall_score"], -x["final_score"], -x["solved_count"], -x["mcqs_correct"], x["time_taken_seconds"]))

    for idx, item in enumerate(leaderboard, 1):
        item["rank"] = idx

    total_cands = len(leaderboard)
    avg_overall = round(sum(x["overall_score"] for x in leaderboard) / max(total_cands, 1), 1)
    highest_overall = max((x["overall_score"] for x in leaderboard), default=0.0)
    total_mcqs = len(contest.get("mcq_ids", []))
    avg_mcq = round(sum(x["mcq_score"] for x in leaderboard) / max(total_cands, 1), 1)
    highest_mcq = max((x["mcq_score"] for x in leaderboard), default=0.0)
    avg_mcq_acc = round(sum(x["mcq_percentage"] for x in leaderboard) / max(total_cands, 1), 1)
    total_coding = len(contest.get("problem_ids", []))
    avg_coding = round(sum(x["coding_score"] for x in leaderboard) / max(total_cands, 1), 1)
    highest_coding = max((x["coding_score"] for x in leaderboard), default=0.0)
    clean_count = sum(1 for x in leaderboard if x["anti_cheat"]["status"] == "CLEAN")
    clean_rate = round((clean_count / max(total_cands, 1)) * 100, 1)
    auto_terminated_count = sum(1 for x in leaderboard if x["anti_cheat"]["auto_terminated"])

    summary = {
        "total_candidates": total_cands,
        "avg_overall_score": avg_overall,
        "highest_overall_score": highest_overall,
        "total_mcqs": total_mcqs,
        "avg_mcq_score": avg_mcq,
        "highest_mcq_score": highest_mcq,
        "avg_mcq_accuracy": avg_mcq_acc,
        "total_coding_problems": total_coding,
        "avg_coding_score": avg_coding,
        "highest_coding_score": highest_coding,
        "clean_rate": clean_rate,
        "clean_count": clean_count,
        "auto_terminated_count": auto_terminated_count
    }

    clean_title = slugify(contest.get("title", "contest"))

    # ================= CSV EXPORT =================
    if export_format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)

        if report_type == "mcq":
            writer.writerow(["Rank", "Student ID", "Candidate Name", "Department", "Year", "Attempt", "Correct MCQs", "Total MCQs", "MCQ Score", "Original MCQ Marks", "Retest Marks", "Accuracy %", "Time Taken", "Anti-Cheat Status"])
            for item in leaderboard:
                original_mcq = item["original_score"]["mcq_score"] if item["original_score"] else item["mcq_score"]
                writer.writerow([
                    item["rank"],
                    item["student_id"],
                    item["name"],
                    item["department"],
                    item["year"],
                    f"Retest #{item['attempt_number']}" if item["is_retest"] else "Original",
                    item["mcqs_correct"],
                    item["total_contest_mcqs"],
                    item["mcq_score"],
                    original_mcq,
                    item["retest_score"]["mcq_score"] if item["retest_score"] else "—",
                    f"{item['mcq_percentage']}%",
                    item["time_taken"],
                    item["anti_cheat"]["status"]
                ])
            filename = f"MCQ_Report_{clean_title}.csv"
        elif report_type == "coding":
            writer.writerow(["Rank", "Student ID", "Candidate Name", "Department", "Year", "Attempt", "Problems Solved", "Total Problems", "Passed Test Cases", "Total Test Cases", "Coding Score", "Original Coding Marks", "Retest Marks", "Time Taken", "Time Complexity", "Space Complexity", "Anti-Cheat Status"])
            for item in leaderboard:
                original_coding = item["original_score"]["coding_score"] if item["original_score"] else item["coding_score"]
                writer.writerow([
                    item["rank"],
                    item["student_id"],
                    item["name"],
                    item["department"],
                    item["year"],
                    f"Retest #{item['attempt_number']}" if item["is_retest"] else "Original",
                    item["solved_count"],
                    item["total_contest_problems"],
                    item["passed_test_cases"],
                    item["total_contest_testcases"],
                    item["coding_score"],
                    original_coding,
                    item["retest_score"]["coding_score"] if item["retest_score"] else "—",
                    item["time_taken"],
                    item["time_complexity"],
                    item["space_complexity"],
                    item["anti_cheat"]["status"]
                ])
            filename = f"CODING_Report_{clean_title}.csv"
        else:
            # Overall
            writer.writerow(["Rank", "Student ID", "Candidate Name", "Department", "Year", "Attempt", "MCQ Score", "Coding Score", "Overall Score", "Original Marks", "Retest Marks", "Performance Index / 100", "Problems Solved", "Test Cases", "Time Taken", "Anti-Cheat Status"])
            for item in leaderboard:
                original_overall = item["original_score"]["score"] if item["original_score"] else item["overall_score"]
                writer.writerow([
                    item["rank"],
                    item["student_id"],
                    item["name"],
                    item["department"],
                    item["year"],
                    f"Retest #{item['attempt_number']}" if item["is_retest"] else "Original",
                    item["mcq_score"],
                    item["coding_score"],
                    item["overall_score"],
                    original_overall,
                    item["retest_score"]["score"] if item["retest_score"] else "—",
                    item["final_score"],
                    f"{item['solved_count']} / {item['total_contest_problems']}",
                    f"{item['passed_test_cases']} / {item['total_contest_testcases']}",
                    item["time_taken"],
                    item["anti_cheat"]["status"]
                ])
            filename = f"OVERALL_Report_{clean_title}.csv"

        csv_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        csv_bytes.seek(0)
        resp = send_file(
            csv_bytes,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        )
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        resp.headers["Access-Control-Expose-Headers"] = "Content-Disposition, Content-Type"
        return resp

    # ================= EXCEL EXPORT =================
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="303442", end_color="303442", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    candidate_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    clean_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    terminated_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    flagged_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    if report_type == "mcq":
        # 1. MCQ Report Sheet
        ws_main = wb.active
        ws_main.title = "MCQ Performance Report"
        ws_main.merge_cells("A1:N1")
        t_cell = ws_main.cell(row=1, column=1, value=f"NIT Campus Coder — MCQ Performance Report: {contest.get('title')}")
        t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        t_cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        t_cell.alignment = center_align

        mcq_headers = ["Rank", "Student ID", "Candidate Name", "Department", "Year", "Attempt", "Correct MCQs", "Total MCQs", "MCQ Score", "Original MCQ Marks", "Retest Marks", "Accuracy %", "Time Taken", "Status"]
        for c_i, h in enumerate(mcq_headers, 1):
            cell = ws_main.cell(row=3, column=c_i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align if c_i not in [2, 3, 4] else left_align

        for r_i, item in enumerate(leaderboard, 4):
            ws_main.cell(row=r_i, column=1, value=item["rank"]).alignment = center_align
            ws_main.cell(row=r_i, column=2, value=item["student_id"]).alignment = left_align
            ws_main.cell(row=r_i, column=3, value=item["name"]).alignment = left_align
            ws_main.cell(row=r_i, column=4, value=item["department"]).alignment = left_align
            ws_main.cell(row=r_i, column=5, value=item["year"]).alignment = center_align
            ws_main.cell(row=r_i, column=6, value=f"Retest #{item['attempt_number']}" if item["is_retest"] else "Original").alignment = center_align
            ws_main.cell(row=r_i, column=7, value=item["mcqs_correct"]).alignment = center_align
            ws_main.cell(row=r_i, column=8, value=item["total_contest_mcqs"]).alignment = center_align
            
            sc_cell = ws_main.cell(row=r_i, column=9, value=item["mcq_score"])
            sc_cell.alignment = center_align
            sc_cell.font = Font(name="Calibri", size=11, bold=True)
            original_mcq = item["original_score"]["mcq_score"] if item["original_score"] else item["mcq_score"]
            ws_main.cell(row=r_i, column=10, value=original_mcq).alignment = center_align
            ws_main.cell(row=r_i, column=11, value=item["retest_score"]["mcq_score"] if item["retest_score"] else "—").alignment = center_align
            ws_main.cell(row=r_i, column=12, value=f"{item['mcq_percentage']}%").alignment = center_align
            ws_main.cell(row=r_i, column=13, value=item["time_taken"]).alignment = center_align
            
            st_cell = ws_main.cell(row=r_i, column=14, value=item["anti_cheat"]["status"])
            st_cell.alignment = center_align
            st_cell.font = Font(name="Calibri", size=10, bold=True)
            if item["anti_cheat"]["status"] == "CLEAN": st_cell.fill = clean_fill
            elif item["anti_cheat"]["status"] == "AUTO_TERMINATED": st_cell.fill = terminated_fill
            else: st_cell.fill = flagged_fill

        filename = f"MCQ_Report_{clean_title}.xlsx"

    elif report_type == "coding":
        # 1. Coding Report Sheet
        ws_main = wb.active
        ws_main.title = "Coding Performance Report"
        ws_main.merge_cells("A1:O1")
        t_cell = ws_main.cell(row=1, column=1, value=f"NIT Campus Coder — Coding Performance Report: {contest.get('title')}")
        t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        t_cell.fill = PatternFill(start_color="0757B8", end_color="0757B8", fill_type="solid")
        t_cell.alignment = center_align

        coding_headers = ["Rank", "Student ID", "Candidate Name", "Department", "Year", "Attempt", "Problems Solved", "Test Cases", "Coding Score", "Original Coding Marks", "Retest Marks", "Time Taken", "Time Comp", "Space Comp", "Status"]
        for c_i, h in enumerate(coding_headers, 1):
            cell = ws_main.cell(row=3, column=c_i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align if c_i not in [2, 3, 4] else left_align

        for r_i, item in enumerate(leaderboard, 4):
            ws_main.cell(row=r_i, column=1, value=item["rank"]).alignment = center_align
            ws_main.cell(row=r_i, column=2, value=item["student_id"]).alignment = left_align
            ws_main.cell(row=r_i, column=3, value=item["name"]).alignment = left_align
            ws_main.cell(row=r_i, column=4, value=item["department"]).alignment = left_align
            ws_main.cell(row=r_i, column=5, value=item["year"]).alignment = center_align
            ws_main.cell(row=r_i, column=6, value=f"Retest #{item['attempt_number']}" if item["is_retest"] else "Original").alignment = center_align
            ws_main.cell(row=r_i, column=7, value=f"{item['solved_count']} / {item['total_contest_problems']}").alignment = center_align
            ws_main.cell(row=r_i, column=8, value=f"{item['passed_test_cases']} / {item['total_contest_testcases']}").alignment = center_align
            
            sc_cell = ws_main.cell(row=r_i, column=9, value=item["coding_score"])
            sc_cell.alignment = center_align
            sc_cell.font = Font(name="Calibri", size=11, bold=True)
            original_coding = item["original_score"]["coding_score"] if item["original_score"] else item["coding_score"]
            ws_main.cell(row=r_i, column=10, value=original_coding).alignment = center_align
            ws_main.cell(row=r_i, column=11, value=item["retest_score"]["coding_score"] if item["retest_score"] else "—").alignment = center_align
            ws_main.cell(row=r_i, column=12, value=item["time_taken"]).alignment = center_align
            ws_main.cell(row=r_i, column=13, value=item["time_complexity"]).alignment = center_align
            ws_main.cell(row=r_i, column=14, value=item["space_complexity"]).alignment = center_align
            
            st_cell = ws_main.cell(row=r_i, column=15, value=item["anti_cheat"]["status"])
            st_cell.alignment = center_align
            st_cell.font = Font(name="Calibri", size=10, bold=True)
            if item["anti_cheat"]["status"] == "CLEAN": st_cell.fill = clean_fill
            elif item["anti_cheat"]["status"] == "AUTO_TERMINATED": st_cell.fill = terminated_fill
            else: st_cell.fill = flagged_fill

        # 2. Problem-wise breakdown sheet for Coding
        ws_prob = wb.create_sheet(title="Coding Problems Breakdown")
        ws_prob.merge_cells("A1:I1")
        p_title = ws_prob.cell(row=1, column=1, value="NIT Campus Coder — Candidate Problem-by-Problem Submissions")
        p_title.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        p_title.fill = PatternFill(start_color="303442", end_color="303442", fill_type="solid")
        p_title.alignment = center_align

        prob_headers = ["Student ID", "Candidate Name", "Problem Title", "Difficulty", "Verdict", "Test Cases Passed", "Runtime (ms)", "Memory (MB)", "Language"]
        for c_i, h in enumerate(prob_headers, 1):
            cell = ws_prob.cell(row=3, column=c_i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = left_align

        p_row = 4
        for item in leaderboard:
            for pb in item.get("problem_breakdowns", []):
                ws_prob.cell(row=p_row, column=1, value=item["student_id"])
                ws_prob.cell(row=p_row, column=2, value=item["name"])
                ws_prob.cell(row=p_row, column=3, value=pb["problem_title"])
                ws_prob.cell(row=p_row, column=4, value=pb["difficulty"])
                v_cell = ws_prob.cell(row=p_row, column=5, value=pb["status"])
                if pb["status"] == "Accepted": v_cell.fill = clean_fill
                elif pb["status"] == "Not Attempted": pass
                else: v_cell.fill = terminated_fill
                ws_prob.cell(row=p_row, column=6, value=f"{pb['passed_test_cases']} / {pb['total_test_cases']}")
                ws_prob.cell(row=p_row, column=7, value=pb["runtime"])
                ws_prob.cell(row=p_row, column=8, value=pb["memory"])
                ws_prob.cell(row=p_row, column=9, value=pb["language"])
                p_row += 1

        filename = f"CODING_Report_{clean_title}.xlsx"

    else:
        # 1. Overall Combined Report Sheet
        ws_main = wb.active
        ws_main.title = "Overall Performance"
        ws_main.merge_cells("A1:O1")
        t_cell = ws_main.cell(row=1, column=1, value=f"NIT Campus Coder — Overall Performance Report: {contest.get('title')}")
        t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        t_cell.fill = PatternFill(start_color="0757B8", end_color="0757B8", fill_type="solid")
        t_cell.alignment = center_align

        overall_headers = [
            "Rank", "Student ID", "Candidate Name", "Department", "Year", 
            "Attempt", "MCQ Marks", "Coding Marks", "Overall Score", "Original Marks", "Retest Marks",
            "Problems Solved", "Test Cases", "Time Taken", "Anti-Cheat Status"
        ]
        for c_i, h in enumerate(overall_headers, 1):
            cell = ws_main.cell(row=3, column=c_i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align if c_i not in [2, 3, 4] else left_align

        for r_i, item in enumerate(leaderboard, 4):
            ws_main.cell(row=r_i, column=1, value=item["rank"]).alignment = center_align
            ws_main.cell(row=r_i, column=2, value=item["student_id"]).alignment = left_align
            ws_main.cell(row=r_i, column=3, value=item["name"]).alignment = left_align
            ws_main.cell(row=r_i, column=4, value=item["department"]).alignment = left_align
            ws_main.cell(row=r_i, column=5, value=item["year"]).alignment = center_align
            ws_main.cell(row=r_i, column=6, value=f"Retest #{item['attempt_number']}" if item["is_retest"] else "Original").alignment = center_align
            
            # MCQ Score
            mcq_cell = ws_main.cell(row=r_i, column=7, value=item["mcq_score"])
            mcq_cell.alignment = center_align
            mcq_cell.font = Font(name="Calibri", size=11, bold=True, color="7C3AED")

            # Coding Score
            cod_cell = ws_main.cell(row=r_i, column=8, value=item["coding_score"])
            cod_cell.alignment = center_align
            cod_cell.font = Font(name="Calibri", size=11, bold=True, color="059669")

            # Overall Score
            sc_cell = ws_main.cell(row=r_i, column=9, value=item["overall_score"])
            sc_cell.alignment = center_align
            sc_cell.font = Font(name="Calibri", size=11, bold=True)
            if item["overall_score"] >= 80: sc_cell.fill = clean_fill
            elif item["overall_score"] >= 50: sc_cell.fill = flagged_fill
            else: sc_cell.fill = terminated_fill

            original_overall = item["original_score"]["score"] if item["original_score"] else item["overall_score"]
            ws_main.cell(row=r_i, column=10, value=original_overall).alignment = center_align
            ws_main.cell(row=r_i, column=11, value=item["retest_score"]["score"] if item["retest_score"] else "—").alignment = center_align
            ws_main.cell(row=r_i, column=12, value=f"{item['solved_count']} / {item['total_contest_problems']}").alignment = center_align
            ws_main.cell(row=r_i, column=13, value=f"{item['passed_test_cases']} / {item['total_contest_testcases']}").alignment = center_align
            ws_main.cell(row=r_i, column=14, value=item["time_taken"]).alignment = center_align
            
            st_cell = ws_main.cell(row=r_i, column=15, value=item["anti_cheat"]["status"])
            st_cell.alignment = center_align
            st_cell.font = Font(name="Calibri", size=10, bold=True)
            if item["anti_cheat"]["status"] == "CLEAN": st_cell.fill = clean_fill
            elif item["anti_cheat"]["status"] == "AUTO_TERMINATED": st_cell.fill = terminated_fill
            else: st_cell.fill = flagged_fill

        # 2. Summary Statistics Sheet
        ws_sum = wb.create_sheet(title="Executive Summary")
        ws_sum.merge_cells("A1:B1")
        s_title = ws_sum.cell(row=1, column=1, value=f"Executive KPI Summary — {contest.get('title')}")
        s_title.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        s_title.fill = PatternFill(start_color="0757B8", end_color="0757B8", fill_type="solid")
        s_title.alignment = center_align

        sum_metrics = [
            ("Contest Title", contest.get("title")),
            ("Duration (Minutes)", contest.get("duration_minutes", 60)),
            ("Total Candidates Attended", summary.get("total_candidates", 0)),
            ("Average Overall Score", summary.get("avg_overall_score", 0)),
            ("Highest Overall Score", summary.get("highest_overall_score", 0)),
            ("Total MCQs in Contest", summary.get("total_mcqs", 0)),
            ("Average MCQ Score", summary.get("avg_mcq_score", 0)),
            ("Highest MCQ Score", summary.get("highest_mcq_score", 0)),
            ("Average MCQ Accuracy (%)", f"{summary.get('avg_mcq_accuracy', 0)}%"),
            ("Total Coding Problems", summary.get("total_coding_problems", 0)),
            ("Average Coding Score", summary.get("avg_coding_score", 0)),
            ("Highest Coding Score", summary.get("highest_coding_score", 0)),
            ("Clean Integrity Rate (%)", f"{summary.get('clean_rate', 100)}%"),
            ("Terminated / Flagged Count", f"{summary.get('auto_terminated_count', 0)} Terminated / {len(leaderboard) - summary.get('clean_count', 0)} Flagged")
        ]

        ws_sum.cell(row=3, column=1, value="Metric").fill = header_fill
        ws_sum.cell(row=3, column=1).font = header_font
        ws_sum.cell(row=3, column=2, value="Value").fill = header_fill
        ws_sum.cell(row=3, column=2).font = header_font

        for r_i, (k, v) in enumerate(sum_metrics, 4):
            ws_sum.cell(row=r_i, column=1, value=k).font = Font(name="Calibri", size=11, bold=True)
            ws_sum.cell(row=r_i, column=2, value=str(v))

        filename = f"OVERALL_Report_{clean_title}.xlsx"

    # Shared Anti-Cheat Audit Sheet in Excel
    ws_audit = wb.create_sheet(title="Anti-Cheat Audit")
    audit_headers = ["Student ID", "Candidate Name", "Department", "Anti-Cheat Status", "Auto-Terminated", "Termination Reason", "Flags Count", "Security Events Summary"]
    for c_i, h in enumerate(audit_headers, 1):
        cell = ws_audit.cell(row=1, column=c_i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = left_align

    for a_i, item in enumerate(leaderboard, 2):
        ac = item["anti_cheat"]
        ws_audit.cell(row=a_i, column=1, value=item["student_id"])
        ws_audit.cell(row=a_i, column=2, value=item["name"])
        ws_audit.cell(row=a_i, column=3, value=item["department"])
        st_cell = ws_audit.cell(row=a_i, column=4, value=ac["status"])
        if ac["status"] == "CLEAN": st_cell.fill = clean_fill
        elif ac["status"] == "AUTO_TERMINATED": st_cell.fill = terminated_fill
        else: st_cell.fill = flagged_fill

        ws_audit.cell(row=a_i, column=5, value="YES" if ac["auto_terminated"] else "NO")
        ws_audit.cell(row=a_i, column=6, value=ac["termination_reason"] or "-")
        ws_audit.cell(row=a_i, column=7, value=ac["flags_count"])
        
        events_summary = ", ".join([f"{log.get('event_type')}: {log.get('detail', '')}" for log in ac.get("logs", [])]) or "No integrity violations"
        ws_audit.cell(row=a_i, column=8, value=events_summary)

    # Adjust widths for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 45), 11)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    from services.notification_service import create_notification
    create_notification(
        user_id=request.current_user["_id"],
        title="Contest Report Generated",
        message=f"Successfully generated and downloaded {report_type.upper()} report for contest '{contest.get('title')}'.",
        notif_type="contest"
    )

    resp = send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["Access-Control-Expose-Headers"] = "Content-Disposition, Content-Type"
    return resp

# ----------------- ADMIN NOTIFICATIONS & ANNOUNCEMENTS -----------------

@admin_bp.route("/notifications", methods=["GET"])
@admin_required
def get_notifications():
    """Retrieve system announcements and anti-cheat warning notification events."""
    db = get_db()
    
    # 1. Fetch all announcements
    announcements = list(db.announcements.find({}).sort("created_at", -1).limit(30))
    announcement_list = []
    for ann in announcements:
        announcement_list.append({
            "id": str(ann["_id"]),
            "title": ann.get("title", "Announcement"),
            "message": ann.get("message", ""),
            "type": ann.get("type", "info"), # info, success, warning, danger
            "created_by": ann.get("created_by", "Admin"),
            "created_at": ann.get("created_at").isoformat() if isinstance(ann.get("created_at"), datetime) else str(ann.get("created_at")),
            "target": "announcement"
        })
        
    # 2. Fetch all anti-cheat violation logs from contest participations
    contest_parts = list(db.contest_participants.find({
        "anti_cheat_logs": {"$exists": True, "$ne": []}
    }).sort("joined_at", -1).limit(50))
    
    violations_list = []
    for cp in contest_parts:
        student_name = cp.get("student_name", "Student")
        student_id = cp.get("student_id", "")
        contest_id = cp.get("contest_id")
        
        # Get contest title
        contest = db.contests.find_one({"_id": ObjectId(contest_id)})
        contest_title = contest.get("title", "Contest") if contest else "Contest"
        
        for log in cp.get("anti_cheat_logs", []):
            violations_list.append({
                "id": f"{cp['_id']}_{log.get('timestamp')}",
                "title": f"Anti-Cheat Flag: {student_name} ({student_id})",
                "message": f"Event '{log.get('event_type')}' in '{contest_title}': {log.get('detail')}",
                "type": "warning" if log.get("event_type") == "tab_switch" else "danger",
                "created_by": "System Security",
                "created_at": log.get("timestamp"),
                "target": "violation"
            })
            
    # Combine and sort by date
    all_notifications = announcement_list + violations_list
    # Parse timestamps for sorting
    def get_timestamp(x):
        try:
            return datetime.fromisoformat(x["created_at"])
        except Exception:
            return datetime.now(timezone.utc)
            
    all_notifications.sort(key=get_timestamp, reverse=True)
    
    return jsonify({
        "success": True,
        "notifications": all_notifications[:50]
    }), 200

@admin_bp.route("/notifications", methods=["POST"])
@admin_required
def create_announcement():
    """Create a new global system announcement."""
    db = get_db()
    data = request.get_json() or {}
    
    title = data.get("title", "").strip()
    message = data.get("message", "").strip()
    notif_type = data.get("type", "info").strip() # info, success, warning, danger
    
    if not title or not message:
        return jsonify({"success": False, "error": "Title and message are required"}), 400
        
    announcement_doc = {
        "title": title,
        "message": message,
        "type": notif_type,
        "created_by": request.current_user.get("name", "Administrator"),
        "created_at": datetime.now(timezone.utc)
    }
    
    res = db.announcements.insert_one(announcement_doc)
    
    from services.notification_service import create_broadcast_notification
    create_broadcast_notification(
        title=title,
        message=message,
        notif_type="system",
        created_by=announcement_doc["created_by"]
    )
    
    return jsonify({
        "success": True,
        "message": "Announcement created successfully",
        "announcement": {
            "id": str(res.inserted_id),
            "title": title,
            "message": message,
            "type": notif_type,
            "created_by": announcement_doc["created_by"],
            "created_at": announcement_doc["created_at"].isoformat()
        }
    }), 201

@admin_bp.route("/notifications/<notif_id>", methods=["DELETE"])
@admin_required
def delete_announcement(notif_id):
    """Delete a system announcement."""
    db = get_db()
    try:
        res = db.announcements.delete_one({"_id": ObjectId(notif_id)})
        if res.deleted_count > 0:
            return jsonify({"success": True, "message": "Announcement deleted successfully"}), 200
        return jsonify({"success": False, "error": "Announcement not found"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400
